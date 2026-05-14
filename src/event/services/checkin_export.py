from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from event.models import CheckIn, CheckInRecord, EventRegistration

HEADERS = (
    "Event",
    "Check-in",
    "Ticket Type",
    "Ticket Code",
    "Attendee Name",
    "Attendee Email",
    "Attendee Organization",
    "Member Title",
    "Member Organization",
    "Checked In",
    "Check-in Station",
    "Checked-in At",
    "Scanned By",
)


def build_checkin_export(check_in: CheckIn) -> bytes:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Check-in Export")
    _set_column_widths(worksheet)
    worksheet.append(_header_cells(worksheet))

    records = _records_by_registration(check_in)
    registrations = (
        EventRegistration.objects.filter(event=check_in.event)
        .select_related("event", "ticket", "member")
        .order_by("attendee_first_name", "attendee_last_name", "attendee_email")
    )
    for registration in registrations:
        worksheet.append(_registration_row(check_in, registration, records.get(registration.pk)))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _set_column_widths(worksheet):
    for index, header in enumerate(HEADERS, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(len(header) + 4, 16)


def _header_cells(worksheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cells = []
    for header in HEADERS:
        cell = WriteOnlyCell(worksheet, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cells.append(cell)
    return cells


def _records_by_registration(check_in: CheckIn) -> dict:
    records = (
        CheckInRecord.objects.filter(registration__event=check_in.event)
        .select_related("check_in", "scanned_by")
        .order_by("-created_at")
    )
    return {record.registration_id: record for record in records}


def _registration_row(check_in: CheckIn, registration: EventRegistration, record: CheckInRecord | None):
    member = registration.member
    return [
        check_in.event.name,
        check_in.name,
        registration.ticket.name if registration.ticket_id else "",
        registration.ticket_code,
        registration.attendee_name,
        registration.attendee_email,
        registration.attendee_organization or "",
        member.title or "",
        member.organization or "",
        "Yes" if record else "No",
        record.check_in.name if record else "",
        _format_timestamp(record.created_at) if record else "",
        _scanned_by(record) if record else "",
    ]


def _format_timestamp(value):
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M:%S")


def _scanned_by(record: CheckInRecord) -> str:
    user = record.scanned_by
    if not user:
        return ""
    return user.get_full_name() or user.get_primary_email() or str(user.pk)
