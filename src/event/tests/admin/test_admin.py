from io import BytesIO
from unittest.mock import patch

from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from authn.models import ContactEmail, ContactPhone, Member
from event.admin.registration import EventRegistrationAdmin
from event.models import CheckIn, CheckInRecord, Event, EventRegistration, Ticket
from event.services import ScheduleSyncStats
from event.tests.helpers import (
    make_event,
    make_member,
    make_registration,
    make_superuser,
    make_ticket,
)


class EventAdminTest(TestCase):
    def setUp(self):
        self.admin_user = make_superuser()
        self.client.login(username="admin@example.com", password="testpass123")

    def test_changelist_accessible(self):
        response = self.client.get("/admin/event/event/")
        self.assertEqual(response.status_code, 200)

    def test_add_form_accessible(self):
        response = self.client.get("/admin/event/event/add/")
        self.assertEqual(response.status_code, 200)

    def test_change_page_shows_inlines_for_staff_without_ticket_model_perms(self):
        """Inlines must not depend on event.add_ticket; match Event admin access instead."""
        editor = Member.objects.create_user(password="testpass123", is_staff=True, is_superuser=False)
        ContactEmail.objects.create(
            member=editor,
            email_address="editor@example.com",
            email_type="primary",
            verified=True,
        )
        ct = ContentType.objects.get_for_model(Event)
        editor.user_permissions.add(Permission.objects.get(content_type=ct, codename="change_event"))
        event = make_event(name="Inline Perm Test")
        self.client.logout()
        self.client.login(username="editor@example.com", password="testpass123")

        response = self.client.get(f"/admin/event/event/{event.pk}/change/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="tickets-group"')
        self.assertContains(response, 'id="questions-group"')

    def test_search_by_name(self):
        make_event(name="Searchable Event")
        response = self.client.get("/admin/event/event/?q=Searchable")
        self.assertEqual(response.status_code, 200)

    def test_list_filter_by_is_live(self):
        response = self.client.get("/admin/event/event/?is_live__exact=1")
        self.assertEqual(response.status_code, 200)

    @patch("event.admin.current_project.sync_schedule")
    def test_pull_schedule_action_triggers_sync(self, mock_sync):
        from event.models import CurrentProjectSchedule

        mock_sync.return_value = ScheduleSyncStats(sections_created=3, tracks_created=3, slots_created=4)
        config = CurrentProjectSchedule.objects.create(name="Demo Day")

        response = self.client.post("/admin/event/currentprojectschedule/pull/")

        self.assertEqual(response.status_code, 302)
        mock_sync.assert_called_once_with(config, sync_type="manual")


class CheckInAdminTest(TestCase):
    def setUp(self):
        self.admin_user = make_superuser(email="checkin-admin@example.com")
        self.client.login(username="checkin-admin@example.com", password="testpass123")
        self.event = make_event(name="Admin Check-in Event")
        self.check_in = CheckIn.objects.create(event=self.event, name="Main Entrance")

    def test_change_page_shows_live_summary_panel(self):
        response = self.client.get(f"/admin/event/checkin/{self.check_in.pk}/change/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "checkin-summary-config")
        self.assertContains(response, "event/css/checkin_change_summary.css")
        self.assertContains(response, "event/js/checkin_change_summary.js")
        self.assertContains(response, "data-checkin-summary")
        self.assertContains(response, "data-summary-total")
        self.assertContains(response, "data-summary-recent-list")
        self.assertContains(response, "Loading recent scans")
        self.assertContains(response, "pollIntervalMs")
        self.assertContains(response, f"/event/check-in/{self.check_in.pk}/status/")
        self.assertContains(response, "Open Check-in Console")
        self.assertContains(response, reverse("admin:event_checkin_scanner", args=[self.check_in.pk]))
        self.assertNotContains(response, "This station")
        self.assertNotContains(response, "data-summary-station")
        self.assertNotContains(response, "Last 5 at this station")
        self.assertNotContains(response, "VIP Gate")
        self.assertNotContains(response, "Check in records")

    def test_changelist_hides_station_scan_count_column(self):
        ticket = make_ticket(self.event, name="General")
        attendee = make_member(email="list-checked-in@example.com", first_name="Ada", last_name="Lovelace")
        registration = make_registration(attendee, self.event, ticket)
        CheckInRecord.objects.create(check_in=self.check_in, registration=registration)

        response = self.client.get("/admin/event/checkin/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Open Console")
        self.assertNotContains(response, "Scans")

    def test_change_page_live_summary_hides_station_count_when_scans_exist(self):
        ticket = make_ticket(self.event, name="General")
        attendee = make_member(email="checked-in@example.com", first_name="Ada", last_name="Lovelace")
        registration = make_registration(attendee, self.event, ticket)
        CheckInRecord.objects.create(check_in=self.check_in, registration=registration)

        response = self.client.get(f"/admin/event/checkin/{self.check_in.pk}/change/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "data-summary-scanned")
        self.assertNotContains(response, "This station")
        self.assertNotContains(response, "data-summary-station")

    def test_scanner_page_is_available_to_staff(self):
        response = self.client.get(reverse("admin:event_checkin_scanner", args=[self.check_in.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "checkin-console-config")
        self.assertContains(response, "event/css/checkin_console.css")
        self.assertContains(response, "event/js/checkin_console.js")
        self.assertContains(response, "data-sync-status")
        self.assertContains(response, "statusPollIntervalMs")
        self.assertContains(response, "data-camera-message")
        self.assertContains(response, "Request camera access")
        self.assertNotContains(response, "This station")
        self.assertNotContains(response, "data-stat-station")
        self.assertContains(response, f"/event/check-in/{self.check_in.pk}/scan/")
        self.assertContains(response, f"/event/check-in/{self.check_in.pk}/status/")
        self.assertContains(response, f"/event/check-in/{self.check_in.pk}/records/__record_id__/undo/")
        self.assertNotContains(response, "i2g-checkin-state")
        self.assertNotContains(response, "function startCamera")

    def test_scanner_page_rejects_non_staff_user(self):
        user = make_member(email="nonstaff-checkin@example.com")
        self.client.force_login(user)

        response = self.client.get(reverse("admin:event_checkin_scanner", args=[self.check_in.pk]))

        self.assertNotEqual(response.status_code, 200)


class EventRegistrationAdminTest(TestCase):
    def setUp(self):
        self.admin_user = make_superuser()
        self.client.login(username="admin@example.com", password="testpass123")

    def test_changelist_accessible(self):
        response = self.client.get("/admin/event/eventregistration/")
        self.assertEqual(response.status_code, 200)

    def test_changelist_shows_send_all_ticket_emails_button(self):
        response = self.client.get("/admin/event/eventregistration/")
        self.assertContains(response, "Send All Tickets")
        self.assertContains(response, reverse("admin:event_eventregistration_send_all_ticket_emails"))

    def test_export_column_picker_includes_member_information(self):
        event = make_event()
        ticket = make_ticket(event)
        member = make_member(email="export-picker@example.com")
        registration = make_registration(member, event, ticket)

        response = self.client.post(
            "/admin/event/eventregistration/",
            {
                "action": "export_data",
                ACTION_CHECKBOX_NAME: str(registration.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Member Title")
        self.assertContains(response, "Member Organization")
        self.assertContains(response, "Member Primary Email")
        self.assertContains(response, "Member Phone Numbers")

    def test_export_excel_includes_member_profile_and_contacts(self):
        event = make_event(name="Export Event")
        ticket = make_ticket(event, name="VIP")
        member = make_member(
            email="export-primary@example.com",
            first_name="Ada",
            middle_name="Byron",
            last_name="Lovelace",
            organization="Analytical Engines",
            title="Chief Scientist",
        )
        ContactEmail.objects.create(
            member=member,
            email_address="export-secondary@example.com",
            email_type="secondary",
            verified=True,
        )
        ContactEmail.objects.create(
            member=member,
            email_address="export-other@example.com",
            email_type="other",
            verified=False,
        )
        ContactPhone.objects.create(
            member=member,
            phone_number="2095551212",
            region="1-US",
            verified=True,
        )
        registration = make_registration(
            member,
            event,
            ticket,
            attendee_first_name="Grace",
            attendee_last_name="Hopper",
            attendee_email="grace@example.com",
            attendee_secondary_email="grace.secondary@example.com",
            attendee_phone="2095552323",
            attendee_organization="Navy",
        )

        response = self.client.post(
            "/admin/event/eventregistration/",
            {
                "action": "export_data",
                ACTION_CHECKBOX_NAME: str(registration.pk),
                "export_confirm": "1",
                "export_format": "xlsx",
                "export_filename": "registration_export",
                "export_fields": [
                    "event_name",
                    "ticket_name",
                    "attendee_email",
                    "attendee_organization",
                    "member_full_name",
                    "member_title",
                    "member_organization",
                    "member_primary_email",
                    "member_secondary_emails",
                    "member_other_emails",
                    "member_phone_numbers",
                ],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(
            rows[0],
            (
                "Event Name",
                "Ticket",
                "Attendee Email",
                "Attendee Organization",
                "Member Full Name",
                "Member Title",
                "Member Organization",
                "Member Primary Email",
                "Member Secondary Emails",
                "Member Other Emails",
                "Member Phone Numbers",
            ),
        )
        self.assertEqual(
            rows[1],
            (
                "Export Event",
                "VIP",
                "grace@example.com",
                "Navy",
                "Ada Byron Lovelace",
                "Chief Scientist",
                "Analytical Engines",
                "export-primary@example.com",
                "export-secondary@example.com",
                "export-other@example.com",
                "(209)555-1212",
            ),
        )

    def test_send_all_ticket_emails_confirmation_page(self):
        event = make_event()
        ticket = make_ticket(event)
        member = make_member(email="ticket-confirm@example.com")
        make_registration(member, event, ticket)

        response = self.client.get(reverse("admin:event_eventregistration_send_all_ticket_emails"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Send Ticket Emails to All Registrants")
        self.assertContains(response, "You are about to send ticket emails to 1 registrant")

    @patch("event.services.ticket_mail.send_ticket_email")
    def test_send_all_ticket_emails_posts_all_registrations(self, mock_send):
        event = make_event()
        ticket = make_ticket(event)
        member_one = make_member(email="ticket-one@example.com")
        member_two = make_member(email="ticket-two@example.com")
        registration_one = make_registration(member_one, event, ticket)
        registration_two = make_registration(member_two, event, ticket)

        response = self.client.post(reverse("admin:event_eventregistration_send_all_ticket_emails"))

        self.assertRedirects(response, reverse("admin:event_eventregistration_changelist"))
        self.assertEqual(mock_send.call_count, 2)
        self.assertEqual(
            {call.args[0].pk for call in mock_send.call_args_list},
            {registration_one.pk, registration_two.pk},
        )

    @patch("event.services.ticket_mail.send_ticket_email")
    def test_send_all_ticket_emails_empty_queryset_does_not_send(self, mock_send):
        response = self.client.post(reverse("admin:event_eventregistration_send_all_ticket_emails"))

        self.assertRedirects(response, reverse("admin:event_eventregistration_changelist"))
        mock_send.assert_not_called()

    def test_has_add_permission(self):
        from django.test import RequestFactory

        factory = RequestFactory()
        request = factory.get("/admin/")
        request.user = self.admin_user
        admin_instance = EventRegistrationAdmin(EventRegistration, None)
        self.assertTrue(admin_instance.has_add_permission(request))

    def test_add_form_accessible(self):
        response = self.client.get("/admin/event/eventregistration/add/")
        self.assertEqual(response.status_code, 200)

    @patch("event.services.registration_sheet_sync.schedule_registration_sync")
    def test_admin_add_creates_registration(self, mock_sync):
        member = Member.objects.create_user(password="testpass123")
        ContactEmail.objects.create(
            member=member,
            email_address="reg@e.com",
            email_type="primary",
            verified=True,
        )
        event = make_event()
        ticket = Ticket.objects.create(event=event, name="GA")
        response = self.client.post(
            "/admin/event/eventregistration/add/",
            {
                "member": str(member.pk),
                "event": str(event.pk),
                "ticket": str(ticket.pk),
                "attendee_first_name": "",
                "attendee_last_name": "",
                "attendee_email": "",
                "attendee_secondary_email": "",
                "attendee_phone": "",
                "attendee_organization": "",
                "phone_verified": "",
                "question_answers": "[]",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(EventRegistration.objects.filter(member=member, event=event).exists())
        mock_sync.assert_called_once_with(event)

    def test_admin_add_rejects_duplicate(self):
        member = Member.objects.create_user(password="testpass123")
        ContactEmail.objects.create(
            member=member,
            email_address="dup@e.com",
            email_type="primary",
            verified=True,
        )
        event = make_event()
        ticket = Ticket.objects.create(event=event, name="GA")
        EventRegistration.objects.create(member=member, event=event, ticket=ticket)
        response = self.client.post(
            "/admin/event/eventregistration/add/",
            {
                "member": str(member.pk),
                "event": str(event.pk),
                "ticket": str(ticket.pk),
                "attendee_first_name": "",
                "attendee_last_name": "",
                "attendee_email": "",
                "attendee_secondary_email": "",
                "attendee_phone": "",
                "attendee_organization": "",
                "phone_verified": "",
                "question_answers": "[]",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(EventRegistration.objects.filter(member=member, event=event).count(), 1)

    def test_admin_add_rejects_mismatched_ticket(self):
        member = Member.objects.create_user(password="testpass123")
        ContactEmail.objects.create(
            member=member,
            email_address="mis@e.com",
            email_type="primary",
            verified=True,
        )
        event1 = make_event(name="Event 1")
        event2 = make_event(name="Event 2")
        ticket_from_event2 = Ticket.objects.create(event=event2, name="VIP")
        response = self.client.post(
            "/admin/event/eventregistration/add/",
            {
                "member": str(member.pk),
                "event": str(event1.pk),
                "ticket": str(ticket_from_event2.pk),
                "attendee_first_name": "",
                "attendee_last_name": "",
                "attendee_email": "",
                "attendee_secondary_email": "",
                "attendee_phone": "",
                "attendee_organization": "",
                "phone_verified": "",
                "question_answers": "[]",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(EventRegistration.objects.filter(member=member, event=event1).exists())

    def test_has_change_permission(self):
        from django.test import RequestFactory

        factory = RequestFactory()
        request = factory.get("/admin/")
        request.user = self.admin_user
        admin_instance = EventRegistrationAdmin(EventRegistration, None)
        self.assertTrue(admin_instance.has_change_permission(request))

    def test_change_page_keeps_only_ticket_editable_in_ticket_section(self):
        from django.test import RequestFactory

        event = make_event()
        ticket = make_ticket(event)
        member = make_member(email="readonly-ticket-section@example.com")
        registration = make_registration(member, event, ticket)
        factory = RequestFactory()
        request = factory.get("/admin/")
        request.user = self.admin_user
        admin_instance = EventRegistrationAdmin(EventRegistration, None)

        readonly_fields = admin_instance.get_readonly_fields(request, obj=registration)

        self.assertNotIn("ticket", readonly_fields)
        self.assertIn("event", readonly_fields)
        self.assertIn("member", readonly_fields)
        self.assertIn("ticket_code", readonly_fields)
        self.assertIn("attendee_email", readonly_fields)
        self.assertIn("send_ticket_email_action", readonly_fields)

    def test_change_form_limits_ticket_choices_to_registration_event(self):
        event = make_event(name="Ticket Choice Event")
        current_ticket = make_ticket(event, name="Current Ticket")
        new_ticket = make_ticket(event, name="Upgrade Ticket")
        other_event = make_event(name="Other Ticket Event")
        other_ticket = make_ticket(other_event, name="Other Event Ticket")
        member = make_member(email="ticket-choice@example.com")
        registration = make_registration(member, event, current_ticket)

        response = self.client.get(f"/admin/event/eventregistration/{registration.pk}/change/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="ticket"')
        self.assertContains(response, 'name="_send_ticket_email"')
        self.assertContains(response, "Send ticket email now")
        self.assertNotContains(response, 'name="send_ticket_email"')
        self.assertNotContains(response, "If checked, a confirmation email")
        self.assertContains(response, str(new_ticket))
        self.assertNotContains(response, str(other_ticket))

    @patch("event.services.registration_sheet_sync.schedule_registration_sync")
    def test_admin_change_updates_registration_ticket(self, mock_sync):
        event = make_event()
        current_ticket = make_ticket(event, name="General")
        new_ticket = make_ticket(event, name="VIP")
        member = make_member(email="ticket-change@example.com")
        registration = make_registration(member, event, current_ticket)

        response = self.client.post(
            f"/admin/event/eventregistration/{registration.pk}/change/",
            {
                "ticket": str(new_ticket.pk),
                "_save": "Save",
            },
        )

        self.assertEqual(response.status_code, 302)
        registration.refresh_from_db()
        self.assertEqual(registration.ticket, new_ticket)
        mock_sync.assert_called_once_with(event)

    @patch("event.services.registration_sheet_sync.schedule_registration_sync")
    def test_admin_change_rejects_ticket_from_other_event(self, mock_sync):
        event = make_event(name="Original Event")
        current_ticket = make_ticket(event, name="General")
        other_event = make_event(name="Other Event")
        other_ticket = make_ticket(other_event, name="Other VIP")
        member = make_member(email="ticket-mismatch-change@example.com")
        registration = make_registration(member, event, current_ticket)

        response = self.client.post(
            f"/admin/event/eventregistration/{registration.pk}/change/",
            {
                "ticket": str(other_ticket.pk),
                "_save": "Save",
            },
        )

        self.assertEqual(response.status_code, 200)
        registration.refresh_from_db()
        self.assertEqual(registration.ticket, current_ticket)
        mock_sync.assert_not_called()

    @patch("event.services.ticket_mail.send_ticket_email")
    @patch("event.services.registration_sheet_sync.schedule_registration_sync")
    def test_admin_change_send_ticket_button_saves_ticket_and_sends_email(self, mock_sync, mock_send):
        event = make_event()
        current_ticket = make_ticket(event, name="General")
        new_ticket = make_ticket(event, name="VIP")
        member = make_member(email="ticket-send-button@example.com")
        registration = make_registration(member, event, current_ticket)

        response = self.client.post(
            f"/admin/event/eventregistration/{registration.pk}/change/",
            {
                "ticket": str(new_ticket.pk),
                "_send_ticket_email": "1",
            },
        )

        self.assertRedirects(response, f"/admin/event/eventregistration/{registration.pk}/change/")
        registration.refresh_from_db()
        self.assertEqual(registration.ticket, new_ticket)
        mock_sync.assert_called_once_with(event)
        mock_send.assert_called_once()
        self.assertEqual(mock_send.call_args.args[0].pk, registration.pk)
        self.assertEqual(mock_send.call_args.args[0].ticket_id, new_ticket.pk)

    def test_has_delete_permission(self):
        from django.test import RequestFactory

        factory = RequestFactory()
        request = factory.get("/admin/")
        request.user = self.admin_user
        admin_instance = EventRegistrationAdmin(EventRegistration, None)
        self.assertTrue(admin_instance.has_delete_permission(request))

    def test_search_by_attendee_fields(self):
        member = Member.objects.create_user(password="testpass123")
        ContactEmail.objects.create(member=member, email_address="u@e.com", email_type="primary", verified=True)
        event = make_event()
        ticket = Ticket.objects.create(event=event, name="GA")
        EventRegistration.objects.create(
            member=member,
            event=event,
            ticket=ticket,
            attendee_first_name="Searchable",
            attendee_last_name="Name",
        )
        response = self.client.get("/admin/event/eventregistration/?q=Searchable")
        self.assertEqual(response.status_code, 200)

    def test_member_info_endpoint(self):
        member = Member.objects.create_user(
            password="testpass123",
            first_name="Ada",
            last_name="Lovelace",
            organization="UCM",
            title="Prof",
        )
        ContactEmail.objects.create(
            member=member,
            email_address="ada@e.com",
            email_type="primary",
            verified=True,
        )
        ContactEmail.objects.create(
            member=member,
            email_address="ada2@e.com",
            email_type="secondary",
            verified=False,
        )
        response = self.client.get(f"/admin/event/eventregistration/member-info/{member.pk}/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["name"], "Ada Lovelace")
        self.assertIn("ada@e.com", data["emails"])
        self.assertIn("ada2@e.com", data["emails"])
        self.assertEqual(data["organization"], "UCM")
        self.assertEqual(data["title"], "Prof")

    def test_event_info_endpoint(self):
        event = make_event(name="Info Test")
        Ticket.objects.create(event=event, name="GA")
        Ticket.objects.create(event=event, name="VIP")
        response = self.client.get(f"/admin/event/eventregistration/event-info/{event.pk}/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["name"], "Info Test")
        self.assertEqual(data["date"], "2025-06-15")
        self.assertEqual(data["location"], "Test Venue")
        self.assertEqual(data["total_registrations"], 0)
        ticket_names = [t["name"] for t in data["tickets"]]
        self.assertIn("GA", ticket_names)
        self.assertIn("VIP", ticket_names)
