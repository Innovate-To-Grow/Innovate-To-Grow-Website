"""
Service for importing members from Excel files.

Expected Excel layout:
    First Name, Last Name, When Started, Last Updated,
    Primary Email, Primary Verified, Primary Subscribed,
    Secondary Email, Secondary Verified, Secondary Subscribed,
    Secondary Expired, Secondary Bounced,
    Phone Number, Phone number subscribed, Phone number verified,
    Organization (optional)
"""

import secrets
import string
from dataclasses import dataclass, field
from datetime import datetime

from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.utils import timezone

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from ..models import ContactEmail, ContactPhone, Member


@dataclass
class ImportResult:
    """Result of a member import operation."""

    success: bool
    created_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    errors: list[str] = field(default_factory=list)
    details: list[dict] = field(default_factory=list)


def generate_random_password(length: int = 12) -> str:
    """Generate a random password."""
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    return "".join(secrets.choice(alphabet) for _ in range(length))


# Column header normalization map
HEADER_MAP = {
    "first name": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "last name": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "when started": "when_started",
    "when_started": "when_started",
    "date joined": "when_started",
    "date_joined": "when_started",
    "last updated": "last_updated",
    "last_updated": "last_updated",
    "primary email": "primary_email",
    "primary_email": "primary_email",
    "email": "primary_email",
    "primary verified": "primary_verified",
    "primary_verified": "primary_verified",
    "primary subscribed": "primary_subscribed",
    "primary_subscribed": "primary_subscribed",
    "secondary email": "secondary_email",
    "secondary_email": "secondary_email",
    "secondary verified": "secondary_verified",
    "secondary_verified": "secondary_verified",
    "secondary subscribed": "secondary_subscribed",
    "secondary_subscribed": "secondary_subscribed",
    "secondary expired": "secondary_expired",
    "secondary_expired": "secondary_expired",
    "secondary bounced": "secondary_bounced",
    "secondary_bounced": "secondary_bounced",
    "phone number": "phone_number",
    "phone_number": "phone_number",
    "phone": "phone_number",
    "phone number subscribed": "phone_subscribed",
    "phone_number_subscribed": "phone_subscribed",
    "phone number verified": "phone_verified",
    "phone_number_verified": "phone_verified",
    "organization": "organization",
    "organization (optional)": "organization",
    "company": "organization",
    "org": "organization",
}


def normalize_header(header: str) -> str:
    """Normalize column header to standard field name."""
    if not header:
        return ""
    header = str(header).strip().lower()
    return HEADER_MAP.get(header, header)


def parse_boolean(value) -> bool:
    """Parse various boolean representations."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    str_val = str(value).strip().lower()
    return str_val in ("true", "yes", "1", "y", "active")


# Date formats to try (order matters — more specific first)
_DATE_FORMATS = [
    "%Y-%m-%d %I:%M %p",  # 2023-03-31 00:01 AM
    "%Y-%m-%d %I:%M:%S %p",  # 2023-03-31 00:01:00 AM
    "%Y-%m-%d %H:%M:%S",  # 2023-03-31 00:01:00
    "%Y-%m-%d %H:%M",  # 2023-03-31 00:01
    "%Y-%m-%d",  # 2023-03-31
    "%m/%d/%Y %I:%M:%S %p",  # 03/31/2023 12:01:00 AM
    "%m/%d/%Y %H:%M:%S",  # 03/31/2023 00:01:00
    "%m/%d/%Y",  # 03/31/2023
    "%m/%d/%y",  # 03/31/23
]


def parse_date(value):
    """Parse date value from Excel cell. Returns timezone-aware datetime or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        if timezone.is_naive(value):
            return timezone.make_aware(value)
        return value
    str_val = str(value).strip()
    if not str_val:
        return None
    for fmt in _DATE_FORMATS:
        try:
            dt = datetime.strptime(str_val, fmt)
            return timezone.make_aware(dt)
        except ValueError:
            continue
    return None


def _clean_phone(value) -> str | None:
    """Clean phone number value from Excel (may come as float/int)."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Strip trailing .0 from Excel numeric cells
    if s.endswith(".0"):
        s = s[:-2]
    return s or None


def _generate_unique_username(email: str, taken: set[str]) -> str:
    """Generate a unique username from email, checking against an in-memory set."""
    base = email.split("@")[0].lower()
    base = "".join(c for c in base if c.isalnum() or c in "._-") or "user"

    username = base
    counter = 2
    while username in taken:
        username = f"{base}_{counter}"
        counter += 1
    taken.add(username)
    return username


# ---------------------------------------------------------------------------
# Parse a single row into a dict (pure data, no DB calls)
# ---------------------------------------------------------------------------


def _parse_row(row_num: int, row_data: dict) -> dict:
    """Parse and validate a single Excel row into a clean dict."""
    primary_email = str(row_data.get("primary_email", "")).strip() if row_data.get("primary_email") else None
    secondary_email = str(row_data.get("secondary_email", "")).strip() if row_data.get("secondary_email") else None

    return {
        "row": row_num,
        "primary_email": primary_email,
        "first_name": str(row_data.get("first_name", "")).strip() if row_data.get("first_name") else "",
        "last_name": str(row_data.get("last_name", "")).strip() if row_data.get("last_name") else "",
        "organization": str(row_data.get("organization", "")).strip() if row_data.get("organization") else "",
        "date_joined": parse_date(row_data.get("when_started")),
        "primary_verified": parse_boolean(row_data.get("primary_verified")),
        "primary_subscribed": parse_boolean(row_data.get("primary_subscribed")),
        "secondary_email": secondary_email,
        "secondary_verified": parse_boolean(row_data.get("secondary_verified")),
        "secondary_subscribed": parse_boolean(row_data.get("secondary_subscribed")),
        "phone_number": _clean_phone(row_data.get("phone_number")),
        "phone_subscribed": parse_boolean(row_data.get("phone_subscribed")),
        "phone_verified": parse_boolean(row_data.get("phone_verified")),
    }


# ---------------------------------------------------------------------------
# Main import function (bulk-optimised)
# ---------------------------------------------------------------------------

BATCH_SIZE = 500


def import_members_from_excel(
    file,
    default_password: str | None = None,
    update_existing: bool = False,
) -> ImportResult:
    """
    Import members from an Excel file.

    Optimised for large files (1000+ rows) by using bulk_create and
    pre-loading lookups into memory instead of per-row DB queries.
    """
    if load_workbook is None:
        return ImportResult(success=False, errors=["openpyxl library not installed. Please run: pip install openpyxl"])

    result = ImportResult(success=True)

    try:
        # ------------------------------------------------------------------
        # 1. Parse Excel into memory
        # ------------------------------------------------------------------
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not raw_rows:
            return ImportResult(success=False, errors=["Excel file is empty"])

        headers = [normalize_header(h) for h in raw_rows[0]]
        if "primary_email" not in headers:
            return ImportResult(success=False, errors=["Excel file must contain a 'Primary Email' column"])

        parsed_rows = []
        seen_in_file: set[str] = set()  # dedup within the file itself
        for row_num, row in enumerate(raw_rows[1:], start=2):
            if not any(row):
                continue
            row_data = dict(zip(headers, row, strict=False))
            parsed = _parse_row(row_num, row_data)

            if not parsed["primary_email"]:
                result.skipped_count += 1
                result.errors.append(f"Row {row_num}: Missing primary email")
                continue

            email_key = parsed["primary_email"].lower()
            if email_key in seen_in_file:
                result.skipped_count += 1
                result.errors.append(f"Row {row_num}: Duplicate email in file ({parsed['primary_email']})")
                continue
            seen_in_file.add(email_key)
            parsed_rows.append(parsed)

        if not parsed_rows:
            return result

        # ------------------------------------------------------------------
        # 2. Pre-load existing DB state for fast lookups
        # ------------------------------------------------------------------
        existing_emails = {e.lower() for e in Member.objects.values_list("email", flat=True)}
        taken_usernames = set(Member.all_objects.values_list("username", flat=True))
        existing_contact_emails = {e.lower() for e in ContactEmail.objects.values_list("email_address", flat=True)}
        existing_phones = set(ContactPhone.objects.values_list("phone_number", flat=True))

        # ------------------------------------------------------------------
        # 3. Hash password once (most expensive single operation)
        # ------------------------------------------------------------------
        hashed_pw = make_password(default_password or generate_random_password())

        # ------------------------------------------------------------------
        # 4. Build object lists (no DB calls)
        # ------------------------------------------------------------------
        members_to_create: list[Member] = []
        emails_to_create: list[ContactEmail] = []
        phones_to_create: list[ContactPhone] = []
        rows_to_update: list[dict] = []

        # Track contact addresses claimed by this import to avoid dupes
        claimed_contact_emails: set[str] = set(existing_contact_emails)
        claimed_phones: set[str] = set(existing_phones)

        now = timezone.now()

        for p in parsed_rows:
            email_key = p["primary_email"].lower()

            if email_key in existing_emails:
                if update_existing:
                    rows_to_update.append(p)
                else:
                    result.skipped_count += 1
                    result.errors.append(f"Row {p['row']}: Member with email {p['primary_email']} already exists")
                continue

            # ---- Build Member instance (UUID generated in Python) ----
            username = _generate_unique_username(p["primary_email"], taken_usernames)
            member = Member(
                username=username,
                email=Member.objects.normalize_email(p["primary_email"]),
                password=hashed_pw,
                first_name=p["first_name"],
                last_name=p["last_name"],
                organization=p["organization"],
                email_subscribe=p["primary_subscribed"],
                is_active=True,
                is_active_member=True,
                date_joined=p["date_joined"] or now,
            )
            members_to_create.append(member)
            existing_emails.add(email_key)

            # ---- Primary ContactEmail ----
            if email_key not in claimed_contact_emails:
                emails_to_create.append(
                    ContactEmail(
                        member=member,
                        email_address=p["primary_email"],
                        email_type="primary",
                        verified=p["primary_verified"],
                        subscribe=p["primary_subscribed"],
                    )
                )
                claimed_contact_emails.add(email_key)

            # ---- Secondary ContactEmail ----
            if p["secondary_email"]:
                sec_key = p["secondary_email"].lower()
                if sec_key not in claimed_contact_emails:
                    emails_to_create.append(
                        ContactEmail(
                            member=member,
                            email_address=p["secondary_email"],
                            email_type="secondary",
                            verified=p["secondary_verified"],
                            subscribe=p["secondary_subscribed"],
                        )
                    )
                    claimed_contact_emails.add(sec_key)

            # ---- ContactPhone ----
            if p["phone_number"] and p["phone_number"] not in claimed_phones:
                phones_to_create.append(
                    ContactPhone(
                        member=member,
                        phone_number=p["phone_number"],
                        region="US",
                        subscribe=p["phone_subscribed"],
                        verified=p["phone_verified"],
                    )
                )
                claimed_phones.add(p["phone_number"])

            result.created_count += 1

        # ------------------------------------------------------------------
        # 5. Bulk-insert everything in one transaction
        # ------------------------------------------------------------------
        with transaction.atomic():
            if members_to_create:
                Member.objects.bulk_create(members_to_create, batch_size=BATCH_SIZE)
            if emails_to_create:
                ContactEmail.objects.bulk_create(emails_to_create, batch_size=BATCH_SIZE)
            if phones_to_create:
                ContactPhone.objects.bulk_create(phones_to_create, batch_size=BATCH_SIZE)

        # ------------------------------------------------------------------
        # 6. Handle updates (row-by-row, typically far fewer)
        # ------------------------------------------------------------------
        if rows_to_update:
            _bulk_update_members(rows_to_update, result, claimed_contact_emails, claimed_phones)

    except Exception as e:
        result.success = False
        result.errors.append(f"Error processing file: {str(e)}")

    return result


# ---------------------------------------------------------------------------
# Update path (row-by-row with savepoints)
# ---------------------------------------------------------------------------


def _bulk_update_members(
    rows: list[dict],
    result: ImportResult,
    claimed_contact_emails: set[str],
    claimed_phones: set[str],
):
    """Update existing members. Uses per-row savepoints for safety."""
    # Pre-fetch all members to update in one query
    emails = [r["primary_email"] for r in rows]
    member_map = {m.email.lower(): m for m in Member.objects.filter(email__in=emails)}

    for p in rows:
        member = member_map.get(p["primary_email"].lower())
        if not member:
            result.skipped_count += 1
            continue

        try:
            with transaction.atomic():
                _update_single_member(member, p, claimed_contact_emails, claimed_phones)
            result.updated_count += 1
        except Exception as e:
            result.skipped_count += 1
            result.errors.append(f"Row {p['row']}: {e}")


def _update_single_member(member, p, claimed_contact_emails, claimed_phones):
    """Update one member and their contact info."""
    if p["first_name"]:
        member.first_name = p["first_name"]
    if p["last_name"]:
        member.last_name = p["last_name"]
    if p["organization"]:
        member.organization = p["organization"]
    member.email_subscribe = p["primary_subscribed"]
    member.save()

    # Primary ContactEmail
    email_key = member.email.lower()
    if email_key not in claimed_contact_emails:
        ContactEmail.objects.update_or_create(
            member=member,
            email_address=member.email,
            defaults={"email_type": "primary", "verified": p["primary_verified"], "subscribe": p["primary_subscribed"]},
        )
        claimed_contact_emails.add(email_key)
    else:
        ContactEmail.objects.filter(member=member, email_address=member.email).update(
            verified=p["primary_verified"], subscribe=p["primary_subscribed"]
        )

    # Secondary ContactEmail
    if p["secondary_email"]:
        sec_key = p["secondary_email"].lower()
        member.contact_emails.filter(email_type="secondary").exclude(email_address=p["secondary_email"]).delete()
        if sec_key not in claimed_contact_emails:
            ContactEmail.objects.update_or_create(
                member=member,
                email_address=p["secondary_email"],
                defaults={
                    "email_type": "secondary",
                    "verified": p["secondary_verified"],
                    "subscribe": p["secondary_subscribed"],
                },
            )
            claimed_contact_emails.add(sec_key)
    else:
        member.contact_emails.filter(email_type="secondary").delete()

    # Phone
    if p["phone_number"]:
        if p["phone_number"] not in claimed_phones:
            existing_phone = member.contact_phones.first()
            if existing_phone:
                existing_phone.phone_number = p["phone_number"]
                existing_phone.subscribe = p["phone_subscribed"]
                existing_phone.verified = p["phone_verified"]
                existing_phone.save()
            else:
                ContactPhone.objects.create(
                    member=member,
                    phone_number=p["phone_number"],
                    region="US",
                    subscribe=p["phone_subscribed"],
                    verified=p["phone_verified"],
                )
            claimed_phones.add(p["phone_number"])
    else:
        member.contact_phones.all().delete()


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------


def generate_template_excel() -> bytes:
    """Generate a template Excel file for member import."""
    if load_workbook is None:
        raise ImportError("openpyxl library not installed")

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = [
        ("first_name", "First Name", 15),
        ("last_name", "Last Name", 15),
        ("when_started", "When Started", 15),
        ("last_updated", "Last Updated", 15),
        ("primary_email", "Primary Email", 30),
        ("primary_verified", "Primary Verified", 16),
        ("primary_subscribed", "Primary Subscribed", 18),
        ("secondary_email", "Secondary Email", 30),
        ("secondary_verified", "Secondary Verified", 18),
        ("secondary_subscribed", "Secondary Subscribed", 20),
        ("secondary_expired", "Secondary Expired", 18),
        ("secondary_bounced", "Secondary Bounced", 18),
        ("phone_number", "Phone Number", 18),
        ("phone_subscribed", "Phone number subscribed", 22),
        ("phone_verified", "Phone number verified", 22),
        ("organization", "Organization (optional)", 25),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, (_field, label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    example_data = [
        "John",
        "Doe",
        "2024-01-15",
        "2024-06-01",
        "john.doe@example.com",
        "TRUE",
        "TRUE",
        "john.d@gmail.com",
        "FALSE",
        "FALSE",
        "FALSE",
        "FALSE",
        "+12095551234",
        "TRUE",
        "FALSE",
        "UC Merced",
    ]

    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border

    ws_help = wb.create_sheet(title="Instructions")
    instructions = [
        ("Field Descriptions", ""),
        ("", ""),
        ("First Name", "Member's first name"),
        ("Last Name", "Member's last name"),
        ("When Started", "Date the member joined (YYYY-MM-DD or MM/DD/YYYY)"),
        ("Last Updated", "Last update date (informational only, not imported)"),
        ("Primary Email", "Primary email address (REQUIRED - used as login email)"),
        ("Primary Verified", "Whether the primary email is verified (TRUE/FALSE)"),
        ("Primary Subscribed", "Whether the primary email is subscribed (TRUE/FALSE)"),
        ("Secondary Email", "Secondary/alternate email address (optional)"),
        ("Secondary Verified", "Whether the secondary email is verified (TRUE/FALSE)"),
        ("Secondary Subscribed", "Whether the secondary email is subscribed (TRUE/FALSE)"),
        ("Secondary Expired", "Whether the secondary email has expired (informational only)"),
        ("Secondary Bounced", "Whether the secondary email has bounced (informational only)"),
        ("Phone Number", "Phone number in E.164 format, e.g. +12095551234 (optional)"),
        ("Phone number subscribed", "Whether the phone is subscribed (TRUE/FALSE)"),
        ("Phone number verified", "Whether the phone is verified (TRUE/FALSE)"),
        ("Organization (optional)", "Organization or company name"),
        ("", ""),
        ("Notes", ""),
        ("", "Username is auto-generated from the primary email address"),
        ("", "A random password is generated unless a default is specified"),
        ("", "Existing members (matched by primary email) can be updated if the option is enabled"),
        ("", "'Last Updated', 'Secondary Expired', and 'Secondary Bounced' are informational and not stored"),
    ]

    for row, (col1, col2) in enumerate(instructions, 1):
        ws_help.cell(row=row, column=1, value=col1)
        ws_help.cell(row=row, column=2, value=col2)

    ws_help.column_dimensions["A"].width = 25
    ws_help.column_dimensions["B"].width = 60

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
