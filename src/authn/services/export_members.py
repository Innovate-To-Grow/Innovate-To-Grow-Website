"""
Export members to Excel (.xlsx) using openpyxl in write-only mode for performance.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def export_members_to_excel(queryset) -> bytes:
    """
    Export a queryset of Member objects to a styled Excel workbook.

    Uses write-only mode and prefetch_related for efficient handling of large datasets.
    Columns match the import template for round-trip compatibility.

    Args:
        queryset: A Django QuerySet of Member objects.

    Returns:
        bytes: The Excel file content.
    """
    queryset = queryset.only(
        "id",
        "first_name",
        "last_name",
        "middle_name",
        "title",
        "organization",
        "is_active",
        "is_staff",
        "date_joined",
    ).prefetch_related("contact_emails", "contact_phones")

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Members")

    # Column definitions: (header_label, width)
    columns = [
        ("Member UUID", 38),
        ("First Name", 15),
        ("Last Name", 15),
        ("Middle Name", 15),
        ("Title", 20),
        ("Organization", 25),
        ("Active", 10),
        ("Staff", 10),
        ("When Started", 18),
        ("Primary Email", 30),
        ("Primary Verified", 16),
        ("Primary Subscribed", 18),
        ("Secondary Email", 30),
        ("Secondary Verified", 18),
        ("Secondary Subscribed", 20),
        ("Phone Number", 18),
        ("Phone Subscribed", 18),
        ("Phone Verified", 18),
    ]

    # Set column widths before appending rows
    for col_idx, (_label, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header styles (matching import template)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write styled header row
    header_cells = []
    for label, _width in columns:
        cell = WriteOnlyCell(ws, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        header_cells.append(cell)
    ws.append(header_cells)

    # Write data rows — plain values, no per-cell styling for performance
    for member in queryset:
        contact_emails = member.contact_emails.all()
        contact_phones = member.contact_phones.all()

        primary = next((ce for ce in contact_emails if ce.email_type == "primary"), None)
        secondary = next((ce for ce in contact_emails if ce.email_type == "secondary"), None)
        phone = next(iter(contact_phones), None)

        ws.append(
            [
                str(member.id),
                member.first_name,
                member.last_name,
                member.middle_name or "",
                member.title or "",
                member.organization or "",
                "Yes" if member.is_active else "No",
                "Yes" if member.is_staff else "No",
                member.date_joined.strftime("%Y-%m-%d %H:%M") if member.date_joined else "",
                primary.email_address if primary else "",
                "Yes" if primary and primary.verified else "No",
                "Yes" if primary and primary.subscribe else "No",
                secondary.email_address if secondary else "",
                "Yes" if secondary and secondary.verified else "No",
                "Yes" if secondary and secondary.subscribe else "No",
                phone.phone_number if phone else "",
                "Yes" if phone and phone.subscribe else "No",
                "Yes" if phone and phone.verified else "No",
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
