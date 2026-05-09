"""Download file builders for admin data export."""

import json
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse


def generate_excel_response(admin_obj, request, queryset, columns):
    """Build an .xlsx file and return it as a download response."""
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(admin_obj.model._meta.verbose_name_plural.title())

    for col_idx, (_name, label) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(label) + 4, 14)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    header_cells = []
    for _name, label in columns:
        cell = WriteOnlyCell(ws, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        header_cells.append(cell)
    ws.append(header_cells)

    for obj in queryset:
        row = [admin_obj.get_export_value(obj, name) for name, _label in columns]
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{admin_obj._get_base_filename(request)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def generate_json_response(admin_obj, request, queryset, columns):
    """Build a JSON file exporting selected columns."""
    data = []
    for obj in queryset:
        item = {}
        for name, _label in columns:
            item[name] = admin_obj.get_export_value(obj, name)
        data.append(item)

    body = json.dumps(data, indent=2, ensure_ascii=False, default=str)
    filename = f"{admin_obj._get_base_filename(request)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    response = HttpResponse(body, content_type="application/json")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
