"""Admin mixins for timestamps and data export."""

import json
import uuid
from datetime import date, datetime
from io import BytesIO

from django.contrib import admin
from django.http import HttpResponse
from django.template.response import TemplateResponse

# ---------------------------------------------------------------------------
# Core mixins
# ---------------------------------------------------------------------------


class TimestampedAdminMixin:
    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        for field in ("created_at", "updated_at"):
            if hasattr(self.model, field) and field not in readonly:
                readonly.append(field)
        return readonly

    def get_list_display(self, request):
        list_display = list(super().get_list_display(request))
        if hasattr(self.model, "created_at") and "created_at" not in list_display:
            list_display.append("created_at")
        return list_display


# ---------------------------------------------------------------------------
# Unified data export mixin (Excel + JSON, with column selection)
# ---------------------------------------------------------------------------

_EXPORT_CONFIRM_PARAM = "export_confirm"

_EXPORT_FORMATS = [
    ("xlsx", "Excel (.xlsx)"),
    ("json", "JSON (.json)"),
]


class DataExportMixin:
    """
    Adds a single "Export selected data" admin action with an intermediate
    page that lets users choose the output format (Excel / JSON), a custom
    filename, and which columns to include.

    Customisation hooks on subclasses:
      - export_fields: list of field names to offer (None = all model fields)
      - export_filename: base name for the download (None = model name)
      - get_export_fields(): return list of (field_name, verbose_name) tuples
      - get_export_value(obj, field_name): serialise a single cell value
    """

    export_fields = None
    export_filename = None

    # Backward-compat aliases for the old ExcelExportMixin API
    @property
    def excel_export_fields(self):
        return self.export_fields

    @excel_export_fields.setter
    def excel_export_fields(self, value):
        self.export_fields = value

    @property
    def excel_export_filename(self):
        return self.export_filename

    @excel_export_filename.setter
    def excel_export_filename(self, value):
        self.export_filename = value

    # -- public helpers subclasses may override ---------------------------------

    def get_export_fields(self):
        """Return [(field_name, verbose_name), ...] for every exportable column."""
        if self.export_fields:
            fields = []
            for name in self.export_fields:
                try:
                    field = self.model._meta.get_field(name)
                    fields.append((name, field.verbose_name.title()))
                except Exception:
                    fields.append((name, name.replace("_", " ").title()))
            return fields
        return [(f.name, f.verbose_name.title()) for f in self.model._meta.fields]

    def get_export_value(self, obj, field_name):
        """Serialise a single field value for export."""
        value = getattr(obj, field_name, None)
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, uuid.UUID):
            return str(value)
        if isinstance(value, list | dict):
            return json.dumps(value, ensure_ascii=False, default=str)
        if hasattr(value, "pk"):
            return str(value)
        return value

    # Backward-compat aliases
    def get_excel_export_fields(self):
        return self.get_export_fields()

    def get_excel_export_value(self, obj, field_name):
        return self.get_export_value(obj, field_name)

    # -- admin action -----------------------------------------------------------

    @admin.action(description="Export selected data")
    def export_data(self, request, queryset):
        if _EXPORT_CONFIRM_PARAM not in request.POST:
            return self._render_column_selection(request, queryset)
        fmt = request.POST.get("export_format", "xlsx")
        if fmt == "json":
            return self._generate_json(request, queryset)
        return self._generate_excel(request, queryset)

    def get_actions(self, request):
        actions = super().get_actions(request)
        actions["export_data"] = (
            type(self).export_data,
            "export_data",
            "Export selected data",
        )
        return actions

    # -- internals --------------------------------------------------------------

    def _get_base_filename(self, request):
        return request.POST.get("export_filename", "").strip() or self.export_filename or self.model._meta.model_name

    def _resolve_columns(self, request):
        """Return [(field_name, label), ...] from user selection, or all columns."""
        selected = request.POST.getlist("export_fields")
        all_fields = dict(self.get_export_fields())
        columns = [(name, all_fields[name]) for name in selected if name in all_fields]
        if not columns:
            columns = list(self.get_export_fields())
        return columns

    def _render_column_selection(self, request, queryset):
        """Show the intermediate column / format picker page."""
        available_fields = self.get_export_fields()
        pks = queryset.values_list("pk", flat=True)

        preview_limit = 5
        preview_rows = []
        for obj in queryset[:preview_limit]:
            preview_rows.append([(name, self.get_export_value(obj, name)) for name, _label in available_fields])

        context = {
            **self.admin_site.each_context(request),
            "title": f"Export {self.model._meta.verbose_name_plural.title()}",
            "queryset": queryset,
            "pks": [str(pk) for pk in pks],
            "available_fields": available_fields,
            "preview_rows": preview_rows,
            "formats": _EXPORT_FORMATS,
            "default_filename": self.export_filename or self.model._meta.model_name,
            "opts": self.model._meta,
            "action_checkbox_name": admin.helpers.ACTION_CHECKBOX_NAME,
            "confirm_param": _EXPORT_CONFIRM_PARAM,
            "media": self.media,
        }
        return TemplateResponse(request, "admin/core/export_columns.html", context)

    def _generate_excel(self, request, queryset):
        """Build the .xlsx file and return it as a download response."""
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        columns = self._resolve_columns(request)

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(self.model._meta.verbose_name_plural.title())

        for col_idx, (_name, label) in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(label) + 4, 14)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        header_cells = []
        for _name, label in columns:
            cell = WriteOnlyCell(ws, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            header_cells.append(cell)
        ws.append(header_cells)

        for obj in queryset:
            row = [self.get_export_value(obj, name) for name, _label in columns]
            ws.append(row)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        base_name = self._get_base_filename(request)
        filename = f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generate_json(self, request, queryset):
        """Build a JSON file exporting selected columns."""
        columns = self._resolve_columns(request)

        data = []
        for obj in queryset:
            item = {}
            for name, _label in columns:
                value = self.get_export_value(obj, name)
                item[name] = value
            data.append(item)

        body = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        base_name = self._get_base_filename(request)
        filename = f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        response = HttpResponse(body, content_type="application/json")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# Backward-compat alias
ExcelExportMixin = DataExportMixin
