"""Generate downloadable Excel files for the System Intelligence chat.

Tools call ``create_export`` after the agent has decided which model + filters
the admin wants. The export is built from the same safe-field machinery the
read tools use (no denied apps/models, no sensitive fields, lookups limited to
``SAFE_LOOKUPS``), so the agent cannot exfiltrate data the read tools would
already refuse.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from django.core.files.base import ContentFile
from django.db import models as django_models
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.models.base.system_intelligence import SystemIntelligenceExport
from core.services.system_intelligence_actions.context import current_conversation, current_user_id
from core.services.system_intelligence_actions.exceptions import ActionRequestError
from core.services.system_intelligence_actions.orm import (
    field_output_name,
    resolve_model,
    safe_model_fields,
    validate_query_key,
    validate_selected_fields,
)
from core.services.system_intelligence_actions.utils import json_safe

EXPORT_DEFAULT_LIMIT = 5000
EXPORT_HARD_LIMIT = 25000
MAX_CELL_CHARS = 32_000  # Excel hard cap is 32,767 — leave headroom.


class ExportError(Exception):
    pass


def create_export(
    *,
    app_label: str,
    model_name: str,
    filters: dict[str, Any] | None = None,
    ordering: str | list[str] | None = None,
    fields: list[str] | None = None,
    limit: int | None = None,
    title: str | None = None,
) -> SystemIntelligenceExport:
    """Run a safe query, write rows to an xlsx workbook, persist as an Export.

    Returns the saved ``SystemIntelligenceExport`` instance. Raises
    ``ExportError`` for caller-fixable problems (no rows, model not allowed,
    invalid field/filter, no active chat conversation).
    """
    conversation = current_conversation()
    user_id = current_user_id()
    if not user_id:
        raise ExportError("No active admin user is associated with this export.")

    try:
        model = resolve_model(app_label, model_name, write=False)
    except ActionRequestError as exc:
        raise ExportError(str(exc)) from exc

    safe_fields = safe_model_fields(model, write=False)
    safe_names = {field_output_name(field) for field in safe_fields}
    try:
        selected = validate_selected_fields(fields, safe_names) if fields else sorted(safe_names)
    except ActionRequestError as exc:
        raise ExportError(str(exc)) from exc
    if not selected:
        raise ExportError(f"No exportable fields configured for {model._meta.label}.")

    qs = model.objects.all()
    filters = filters or {}
    if not isinstance(filters, dict):
        raise ExportError("filters must be an object.")
    try:
        for key in filters:
            validate_query_key(key, safe_names)
    except ActionRequestError as exc:
        raise ExportError(str(exc)) from exc
    if filters:
        qs = qs.filter(**filters)

    if ordering:
        ordering_list = [ordering] if isinstance(ordering, str) else list(ordering)
        try:
            for key in ordering_list:
                validate_query_key(key, safe_names)
        except ActionRequestError as exc:
            raise ExportError(str(exc)) from exc
        qs = qs.order_by(*ordering_list)

    capped_limit = min(int(limit or EXPORT_DEFAULT_LIMIT), EXPORT_HARD_LIMIT)
    rows = list(qs.values(*selected)[:capped_limit])
    if not rows:
        raise ExportError("No rows match the requested filters; nothing to export.")

    headers = [_header_label(model, name) for name in selected]
    workbook = _build_workbook(model._meta.label, headers, selected, rows)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = _build_filename(model)
    safe_title = (title or f"{model._meta.verbose_name_plural.title()} export").strip()[:200]
    export = SystemIntelligenceExport(
        conversation=conversation,
        created_by_id=user_id,
        title=safe_title,
        filename=filename,
        model_label=model._meta.label,
        field_names=list(selected),
        row_count=len(rows),
    )
    export.file.save(filename, ContentFile(buffer.getvalue()), save=False)
    export.save()
    return export


def _build_workbook(label: str, headers: list[str], selected: list[str], rows: list[dict]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = label.split(".")[-1][:31] or "Export"
    sheet.append(headers)
    for row in rows:
        sheet.append([_render_cell(row.get(name)) for name in selected])
    _autosize_columns(sheet, headers, rows, selected)
    return workbook


def _render_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        # openpyxl rejects timezone-aware datetimes; persist the local wall-clock
        # value so admins see the same time their admin pages already render.
        if value.tzinfo is not None:
            value = timezone.localtime(value).replace(tzinfo=None)
        return value
    if isinstance(value, bool | int | float):
        return value
    safe = json_safe(value)
    if isinstance(safe, str):
        return safe[:MAX_CELL_CHARS]
    return str(safe)[:MAX_CELL_CHARS]


def _autosize_columns(sheet, headers, rows, selected):
    for index, name in enumerate(selected):
        widest = max(
            (len(str(row.get(name) or "")) for row in rows[:200]),
            default=0,
        )
        width = min(60, max(len(headers[index]) + 2, widest + 2, 12))
        sheet.column_dimensions[get_column_letter(index + 1)].width = width


def _header_label(model: type[django_models.Model], name: str) -> str:
    try:
        field = model._meta.get_field(name)
    except Exception:
        return name.replace("_", " ").title()
    return str(getattr(field, "verbose_name", name)).title()


def _build_filename(model: type[django_models.Model]) -> str:
    stamp = timezone.now().astimezone().strftime("%Y%m%d-%H%M%S")
    slug = model._meta.model_name or model._meta.object_name.lower()
    return f"{slug}-{stamp}.xlsx"
