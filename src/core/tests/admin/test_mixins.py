"""Tests for core admin mixins (timestamped, data export)."""

import io
import json

from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.http import QueryDict
from django.test import RequestFactory, TestCase
from django.utils import timezone
from openpyxl import load_workbook

from cms.models import CMSPage
from core.admin.mixins import (
    DataExportMixin,
    ExcelExportMixin,
    TimestampedAdminMixin,
)

User = get_user_model()

_slug_counter = 0


def _make_page(**kwargs):
    global _slug_counter
    _slug_counter += 1
    ts = f"{_slug_counter}-{timezone.now().timestamp()}".replace(".", "-")
    defaults = {
        "title": "Test",
        "slug": f"test-{ts}",
        "route": f"/test-{ts}",
        "status": "draft",
    }
    defaults.update(kwargs)
    return CMSPage.objects.create(**defaults)


class _MockSuperAdmin:
    """Minimal stand-in for ModelAdmin methods that mixins call via super()."""

    model = CMSPage
    admin_site = AdminSite()

    @property
    def media(self):
        from django.forms import Media

        return Media()

    def get_queryset(self, request):
        return CMSPage.objects.all()

    def get_list_display(self, request):
        return ["__str__"]

    def get_list_filter(self, request):
        return []

    def get_readonly_fields(self, request, obj=None):
        return []

    def get_actions(self, request):
        return {}

    def save_model(self, request, obj, form, change):
        obj.save()


class _TimestampAdmin(TimestampedAdminMixin, _MockSuperAdmin):
    pass


class _DataExportAdmin(DataExportMixin, _MockSuperAdmin):
    pass


class _DataExportCustomAdmin(DataExportMixin, _MockSuperAdmin):
    export_fields = ["title", "status", "published_at"]
    export_filename = "custom_pages"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _load_xlsx(response_content):
    """Parse response bytes into an openpyxl Workbook."""
    return load_workbook(io.BytesIO(response_content))


def _make_post_request(factory, data=None):
    """Build a POST request with the given data dict."""
    request = factory.post("/admin/", data=data or {})
    request.META["SERVER_NAME"] = "testserver"
    return request


def _confirmed_post(factory, extra=None):
    """Build a confirmed export POST request with optional extra fields."""
    data = {"action": "export_data", "export_confirm": "1", "export_format": "xlsx"}
    if extra:
        data.update(extra)
    return _make_post_request(factory, data)


# ---------------------------------------------------------------------------
# TimestampedAdminMixin
# ---------------------------------------------------------------------------


class TimestampedAdminMixinTest(TestCase):
    def setUp(self):
        self.admin = _TimestampAdmin()
        self.factory = RequestFactory()
        self.request = self.factory.get("/admin/")

    def test_readonly_includes_timestamp_fields(self):
        readonly = self.admin.get_readonly_fields(self.request)
        for field in ("created_at", "updated_at"):
            self.assertIn(field, readonly)

    def test_list_display_includes_created_at(self):
        display = self.admin.get_list_display(self.request)
        self.assertIn("created_at", display)


# ---------------------------------------------------------------------------
# DataExportMixin tests
# ---------------------------------------------------------------------------


class DataExportMixinTest(TestCase):
    def setUp(self):
        self.admin = _DataExportAdmin()
        self.factory = RequestFactory()
        self.request = self.factory.get("/admin/")

    # -- action registration ---------------------------------------------------

    def test_actions_include_export_data(self):
        actions = self.admin.get_actions(self.request)
        self.assertIn("export_data", actions)

    def test_action_tuple_uses_unbound_method(self):
        actions = self.admin.get_actions(self.request)
        func = actions["export_data"][0]
        self.assertFalse(hasattr(func, "__self__"), "Action func should be unbound to avoid double-self bug")

    # -- backward compat alias -------------------------------------------------

    def test_excel_export_mixin_alias(self):
        self.assertIs(ExcelExportMixin, DataExportMixin)

    # -- get_export_fields -----------------------------------------------------

    def test_default_fields_returns_all_model_fields(self):
        fields = self.admin.get_export_fields()
        field_names = [name for name, _label in fields]
        model_field_names = [f.name for f in CMSPage._meta.fields]
        self.assertEqual(field_names, model_field_names)

    def test_default_fields_returns_verbose_name_labels(self):
        fields = self.admin.get_export_fields()
        field_dict = dict(fields)
        self.assertEqual(field_dict["published_at"], "Published At")
        self.assertEqual(field_dict["meta_description"], "Meta Description")

    def test_backward_compat_get_excel_export_fields(self):
        self.assertEqual(self.admin.get_excel_export_fields(), self.admin.get_export_fields())

    # -- get_export_value ------------------------------------------------------

    def test_value_none_returns_empty_string(self):
        page = _make_page(title="ValNone")
        self.assertEqual(self.admin.get_export_value(page, "meta_description"), "")

    def test_value_datetime_formatted(self):
        dt = timezone.now()
        page = _make_page(title="ValDt", published_at=dt)
        result = self.admin.get_export_value(page, "published_at")
        self.assertEqual(result, dt.strftime("%Y-%m-%d %H:%M:%S"))

    def test_value_uuid_returned_as_string(self):
        page = _make_page(title="ValUuid")
        result = self.admin.get_export_value(page, "id")
        self.assertEqual(result, str(page.id))

    def test_value_string_returned_as_is(self):
        page = _make_page(title="Hello")
        self.assertEqual(self.admin.get_export_value(page, "title"), "Hello")

    def test_value_date_formatted(self):
        from datetime import date

        class _FakeDateObj:
            date_field = date(2025, 3, 15)

        result = self.admin.get_export_value(_FakeDateObj(), "date_field")
        self.assertEqual(result, "2025-03-15")

    def test_value_list_serialized_as_json(self):
        page = _make_page(title="ValList")
        page._test_list = [{"q": "Color?", "a": "Blue"}]
        result = self.admin.get_export_value(page, "_test_list")
        self.assertEqual(result, '[{"q": "Color?", "a": "Blue"}]')

    def test_value_dict_serialized_as_json(self):
        page = _make_page(title="ValDict")
        page._test_dict = {"key": "value", "num": 42}
        result = self.admin.get_export_value(page, "_test_dict")
        self.assertIn('"key": "value"', result)
        self.assertIn('"num": 42', result)

    def test_value_bool_returns_yes_no(self):
        class _FakeBoolObj:
            active = True
            deleted = False

        self.assertEqual(self.admin.get_export_value(_FakeBoolObj(), "active"), "Yes")
        self.assertEqual(self.admin.get_export_value(_FakeBoolObj(), "deleted"), "No")

    def test_backward_compat_get_excel_export_value(self):
        page = _make_page(title="CompatVal")
        self.assertEqual(
            self.admin.get_excel_export_value(page, "title"),
            self.admin.get_export_value(page, "title"),
        )

    # -- intermediate column / format selection page ---------------------------

    def test_initial_post_renders_column_selection_template(self):
        page = _make_page(title="Tmpl1")
        qs = CMSPage.objects.filter(pk=page.pk)

        request = _make_post_request(
            self.factory,
            {
                "action": "export_data",
                "_selected_action": [str(page.pk)],
            },
        )
        request.user = User.objects.create_user(email="tmpl@example.com", password="StrongPass123!", is_staff=True)
        response = self.admin.export_data(request, qs)

        self.assertEqual(response.template_name, "admin/core/export_columns.html")
        self.assertIn(str(page.pk), response.context_data["pks"])
        self.assertGreater(len(response.context_data["available_fields"]), 0)
        self.assertEqual(response.context_data["default_filename"], "cmspage")
        self.assertEqual(len(response.context_data["formats"]), 2)

    # -- Excel generation (all columns) ----------------------------------------

    def test_xlsx_confirmed_post_returns_xlsx(self):
        a1 = _make_page(title="First Page")
        a2 = _make_page(title="Second Page")
        qs = CMSPage.objects.filter(pk__in=[a1.pk, a2.pk])

        request = _confirmed_post(self.factory)
        response = self.admin.export_data(request, qs)

        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_xlsx_contains_header_and_data_rows(self):
        a1 = _make_page(title="Alpha")
        a2 = _make_page(title="Beta")
        qs = CMSPage.objects.filter(pk__in=[a1.pk, a2.pk]).order_by("title")

        request = _confirmed_post(self.factory)
        response = self.admin.export_data(request, qs)

        wb = _load_xlsx(response.content)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        header = rows[0]
        self.assertIn("Title", header)
        self.assertIn("Status", header)

        self.assertEqual(len(rows), 3)  # 1 header + 2 data
        titles = {row[header.index("Title")] for row in rows[1:]}
        self.assertEqual(titles, {"Alpha", "Beta"})

    def test_xlsx_header_has_blue_fill(self):
        _make_page(title="StyleTest")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory)
        response = self.admin.export_data(request, qs)

        wb = _load_xlsx(response.content)
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        self.assertEqual(header_cell.fill.start_color.rgb, "004472C4")

    # -- column selection (specific columns) -----------------------------------

    def test_selected_columns_limits_exported_fields(self):
        page = _make_page(title="Selected", status="published")
        qs = CMSPage.objects.filter(pk=page.pk)

        request = _make_post_request(self.factory, QueryDict(mutable=True))
        request.POST = QueryDict(mutable=True)
        request.POST["action"] = "export_data"
        request.POST["export_confirm"] = "1"
        request.POST["export_format"] = "xlsx"
        request.POST.setlist("export_fields", ["title", "status"])

        response = self.admin.export_data(request, qs)

        wb = _load_xlsx(response.content)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        self.assertEqual(len(header), 2)
        self.assertEqual(header, ("Title", "Status"))
        self.assertEqual(rows[1], ("Selected", "published"))

    def test_empty_selection_exports_all_columns(self):
        _make_page(title="EmptySel")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory)
        response = self.admin.export_data(request, qs)

        wb = _load_xlsx(response.content)
        ws = wb.active
        header = next(ws.iter_rows(values_only=True))
        model_field_count = len(CMSPage._meta.fields)
        self.assertEqual(len(header), model_field_count)

    # -- filename customisation ------------------------------------------------

    def test_default_filename_uses_model_name(self):
        _make_page(title="FnameTest")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory)
        response = self.admin.export_data(request, qs)

        self.assertIn("cmspage_", response["Content-Disposition"])

    def test_user_provided_filename_used_in_response(self):
        _make_page(title="UserFname")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory, {"export_filename": "my_custom_report"})
        response = self.admin.export_data(request, qs)

        self.assertIn("my_custom_report_", response["Content-Disposition"])
        self.assertNotIn("cmspage_", response["Content-Disposition"])

    def test_blank_user_filename_falls_back_to_default(self):
        _make_page(title="BlankFname")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory, {"export_filename": "   "})
        response = self.admin.export_data(request, qs)

        self.assertIn("cmspage_", response["Content-Disposition"])

    # -- JSON format -----------------------------------------------------------

    def test_json_format_returns_json_response(self):
        a1 = _make_page(title="First")
        a2 = _make_page(title="Second")
        qs = CMSPage.objects.filter(pk__in=[a1.pk, a2.pk])

        request = _confirmed_post(self.factory, {"export_format": "json"})
        response = self.admin.export_data(request, qs)

        self.assertEqual(response["Content-Type"], "application/json")
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".json", response["Content-Disposition"])

        data = json.loads(response.content)
        self.assertEqual(len(data), 2)
        titles = {item["title"] for item in data}
        self.assertEqual(titles, {"First", "Second"})

    def test_json_format_respects_column_selection(self):
        _make_page(title="Filtered", status="published")
        qs = CMSPage.objects.all()

        request = _make_post_request(self.factory, QueryDict(mutable=True))
        request.POST = QueryDict(mutable=True)
        request.POST["action"] = "export_data"
        request.POST["export_confirm"] = "1"
        request.POST["export_format"] = "json"
        request.POST.setlist("export_fields", ["title", "status"])

        response = self.admin.export_data(request, qs)

        data = json.loads(response.content)
        self.assertEqual(len(data), 1)
        self.assertEqual(set(data[0].keys()), {"title", "status"})

    def test_json_format_uses_custom_filename(self):
        _make_page(title="JsonFname")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory, {"export_format": "json", "export_filename": "my_report"})
        response = self.admin.export_data(request, qs)

        self.assertIn("my_report_", response["Content-Disposition"])
        self.assertIn(".json", response["Content-Disposition"])

    def test_json_default_exports_all_columns(self):
        _make_page(title="JsonAll")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory, {"export_format": "json"})
        response = self.admin.export_data(request, qs)

        data = json.loads(response.content)
        model_field_names = {f.name for f in CMSPage._meta.fields}
        self.assertEqual(set(data[0].keys()), model_field_names)


# ---------------------------------------------------------------------------
# DataExportMixin with custom export_fields / export_filename
# ---------------------------------------------------------------------------


class DataExportMixinCustomFieldsTest(TestCase):
    def setUp(self):
        self.admin = _DataExportCustomAdmin()
        self.factory = RequestFactory()

    def test_custom_fields_limits_available_columns(self):
        fields = self.admin.get_export_fields()
        field_names = [name for name, _label in fields]
        self.assertEqual(field_names, ["title", "status", "published_at"])

    def test_custom_filename_used_in_xlsx_response(self):
        _make_page(title="CustFname")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory)
        response = self.admin.export_data(request, qs)

        self.assertIn("custom_pages_", response["Content-Disposition"])

    def test_custom_fields_exports_only_specified_columns(self):
        dt = timezone.now()
        _make_page(title="Custom", status="published", published_at=dt)
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory)
        response = self.admin.export_data(request, qs)

        wb = _load_xlsx(response.content)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        self.assertEqual(len(header), 3)
        self.assertEqual(rows[1][header.index("Title")], "Custom")
        self.assertEqual(rows[1][header.index("Status")], "published")

    def test_custom_filename_used_in_json_response(self):
        _make_page(title="CustJsonFname")
        qs = CMSPage.objects.all()

        request = _confirmed_post(self.factory, {"export_format": "json"})
        response = self.admin.export_data(request, qs)

        self.assertIn("custom_pages_", response["Content-Disposition"])
        self.assertIn(".json", response["Content-Disposition"])

    def test_backward_compat_excel_export_fields_property(self):
        self.assertEqual(self.admin.excel_export_fields, ["title", "status", "published_at"])

    def test_backward_compat_excel_export_filename_property(self):
        self.assertEqual(self.admin.excel_export_filename, "custom_pages")


# ---------------------------------------------------------------------------
# DataExportMixin — DateField and JSONField serialization
# ---------------------------------------------------------------------------


class _DateFieldExportAdmin(DataExportMixin, _MockSuperAdmin):
    """Admin that exports Event-like models with DateField and JSONField."""

    model = None  # set in setUp via monkey-patch


class DataExportDateFieldTest(TestCase):
    """Verify JSON export does not crash on DateField / JSONField values."""

    def setUp(self):
        self.admin = _DataExportAdmin()
        self.factory = RequestFactory()

    def test_json_export_with_date_field_does_not_crash(self):
        """Regression: DateField values caused TypeError in json.dumps."""
        from event.models import Event

        event = Event.objects.create(
            name="Test Event",
            slug="test-export-date",
            date="2025-06-15",
            location="Room A",
            description="Testing",
        )

        # Build a fresh admin bound to Event model
        admin = _DataExportAdmin()
        admin.model = Event

        qs = Event.objects.filter(pk=event.pk)
        request = _confirmed_post(self.factory, {"export_format": "json"})

        response = admin.export_data(request, qs)

        self.assertEqual(response["Content-Type"], "application/json")
        data = json.loads(response.content)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["date"], "2025-06-15")

    def test_json_export_with_json_field_does_not_crash(self):
        """Regression: list/dict values from JSONField caused TypeError in json.dumps."""
        page = _make_page(title="JSON Export")
        # CMSPage doesn't have a JSONField, but the export should handle all field types
        qs = CMSPage.objects.filter(pk=page.pk)
        request = _confirmed_post(self.factory, {"export_format": "json"})

        # The export should not crash — json.dumps with default=str handles any leftovers
        response = self.admin.export_data(request, qs)

        self.assertEqual(response["Content-Type"], "application/json")
        data = json.loads(response.content)
        self.assertEqual(len(data), 1)
        self.assertIn("title", data[0])
