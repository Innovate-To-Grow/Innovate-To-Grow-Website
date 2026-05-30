"""Line-coverage tests for ``services.exports.creator``.

These target the uncovered branches that the broader ``test_exports.py`` suite
does not exercise: the no-user guard, the empty-field guard, the non-dict
filters guard, the ordering validation paths, and the two private cell/header
helpers' fallback branches.
"""

from datetime import datetime
from unittest.mock import patch

from django.test import TestCase

from apps.authn.models import Member
from apps.event.tests.helpers import make_superuser
from apps.system_intelligence.models import ChatConversation
from apps.system_intelligence.services import actions
from apps.system_intelligence.services.exports import ExportError, create_export
from apps.system_intelligence.services.exports.creator import _header_label, _render_cell


class CreateExportGuardTests(TestCase):
    """Cover the early-return / error guards inside ``create_export``."""

    def setUp(self):
        self.admin_user = make_superuser()
        self.conversation = ChatConversation.objects.create(created_by=self.admin_user)
        # Default: full valid context so individual tests can mutate as needed.
        self.context_tokens = actions.set_action_context(str(self.conversation.pk), str(self.admin_user.pk))

    def tearDown(self):
        actions.reset_action_context(self.context_tokens)

    def test_missing_user_id_raises(self):
        # Re-set the context with a conversation but no user id (line 62 guard).
        actions.reset_action_context(self.context_tokens)
        self.context_tokens = actions.set_action_context(str(self.conversation.pk), None)
        with self.assertRaises(ExportError) as cm:
            create_export(app_label="authn", model_name="Member")
        self.assertEqual(str(cm.exception), "No active admin user is associated with this export.")

    def test_no_exportable_fields_raises(self):
        # When the model exposes no safe fields, ``selected`` is empty (line 76).
        with patch(
            "apps.system_intelligence.services.exports.creator.safe_model_fields",
            return_value=[],
        ):
            with self.assertRaises(ExportError) as cm:
                create_export(app_label="authn", model_name="Member")
        self.assertEqual(str(cm.exception), "No exportable fields configured for authn.Member.")

    def test_non_dict_filters_raises(self):
        # filters that are not a mapping must be rejected (line 81).
        with self.assertRaises(ExportError) as cm:
            create_export(app_label="authn", model_name="Member", filters=["not", "a", "dict"])
        self.assertEqual(str(cm.exception), "filters must be an object.")


class CreateExportOrderingTests(TestCase):
    """Cover the ordering normalization + validation branch (lines 90-97)."""

    def setUp(self):
        self.admin_user = make_superuser()
        self.conversation = ChatConversation.objects.create(created_by=self.admin_user)
        self.context_tokens = actions.set_action_context(str(self.conversation.pk), str(self.admin_user.pk))
        # A couple of extra members so ordering produces a deterministic sequence.
        Member.objects.create_user(email="zoe@example.com", password="testpass123", first_name="Zoe")
        Member.objects.create_user(email="amy@example.com", password="testpass123", first_name="Amy")

    def tearDown(self):
        actions.reset_action_context(self.context_tokens)

    def _read_first_data_column(self, export, column_index):
        from openpyxl import load_workbook

        workbook = load_workbook(export.file.path, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header = rows[0]
        values = [row[column_index] for row in rows[1:]]
        return header, values

    def test_ordering_as_string_orders_rows(self):
        # String ordering takes the ``[ordering]`` branch of line 91 and applies
        # ``order_by`` (line 97).
        export = create_export(
            app_label="authn",
            model_name="Member",
            fields=["first_name"],
            ordering="first_name",
        )
        _, values = self._read_first_data_column(export, 0)
        non_empty = [v for v in values if v]
        self.assertEqual(non_empty, sorted(non_empty))

    def test_ordering_as_list_orders_rows_descending(self):
        # List ordering takes the ``list(ordering)`` branch of line 91.
        export = create_export(
            app_label="authn",
            model_name="Member",
            fields=["first_name"],
            ordering=["-first_name"],
        )
        _, values = self._read_first_data_column(export, 0)
        non_empty = [v for v in values if v]
        self.assertEqual(non_empty, sorted(non_empty, reverse=True))

    def test_invalid_ordering_key_raises(self):
        # An ordering key that is not a safe field surfaces as ExportError
        # (lines 95-96).
        with self.assertRaises(ExportError) as cm:
            create_export(
                app_label="authn",
                model_name="Member",
                ordering="definitely_not_a_field",
            )
        self.assertIn("not allowed", str(cm.exception))


class RenderCellTests(TestCase):
    """Cover ``_render_cell`` fallthrough for non-str json-safe values (line 151)."""

    def test_none_returns_empty_string(self):
        self.assertEqual(_render_cell(None), "")

    def test_naive_datetime_passes_through(self):
        value = datetime(2026, 1, 2, 3, 4, 5)
        self.assertEqual(_render_cell(value), value)

    def test_primitive_passes_through(self):
        self.assertEqual(_render_cell(42), 42)
        self.assertEqual(_render_cell(True), True)

    def test_string_is_truncated(self):
        long = "x" * 40_000
        rendered = _render_cell(long)
        self.assertEqual(len(rendered), 32_000)

    def test_list_value_is_stringified(self):
        # A list is JSON-safe but not a ``str``; line 151 stringifies it.
        rendered = _render_cell([1, 2, 3])
        self.assertEqual(rendered, "[1, 2, 3]")

    def test_dict_value_is_stringified(self):
        rendered = _render_cell({"a": 1})
        self.assertEqual(rendered, "{'a': 1}")


class HeaderLabelTests(TestCase):
    """Cover ``_header_label`` fallback when ``get_field`` raises (lines 167-168)."""

    def test_unknown_field_falls_back_to_titleized_name(self):
        label = _header_label(Member, "not_a_real_field")
        self.assertEqual(label, "Not A Real Field")

    def test_known_field_uses_verbose_name(self):
        label = _header_label(Member, "first_name")
        self.assertEqual(label, "First Name")
