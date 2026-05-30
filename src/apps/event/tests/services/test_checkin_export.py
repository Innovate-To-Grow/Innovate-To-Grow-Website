from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from apps.event.models import CheckIn, CheckInRecord
from apps.event.services.checkin_export import _scanned_by, build_checkin_export
from apps.event.tests.helpers import make_event, make_member, make_registration, make_ticket


class CheckinExportScannedByTest(TestCase):
    def setUp(self):
        self.event = make_event(name="Export Service Event")
        self.ticket = make_ticket(self.event, name="GA")
        self.member = make_member(email="export-svc@example.com", first_name="Ada", last_name="Lovelace")
        self.registration = make_registration(self.member, self.event, self.ticket)
        self.check_in = CheckIn.objects.create(event=self.event, name="Main")

    def test_scanned_by_returns_empty_when_no_user(self):
        record = CheckInRecord.objects.create(
            check_in=self.check_in,
            registration=self.registration,
            scanned_by=None,
        )
        self.assertEqual(_scanned_by(record), "")

    def test_export_leaves_scanned_by_blank_for_anonymous_scan(self):
        CheckInRecord.objects.create(
            check_in=self.check_in,
            registration=self.registration,
            scanned_by=None,
        )

        content = build_checkin_export(self.check_in)

        workbook = load_workbook(BytesIO(content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        # Data row: "Scanned By" is the last column and should be blank
        # (openpyxl read_only mode surfaces an empty string as None).
        self.assertIn(rows[1][-1], ("", None))
        self.assertEqual(rows[1][9], "Yes")
