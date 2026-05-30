"""Tests for the member import template generator and parsing helpers."""

from datetime import datetime
from io import BytesIO

from django.test import TestCase
from django.utils import timezone

from apps.authn.services.import_members.parsing import (
    clean_phone,
    generate_random_password,
    normalize_header,
    parse_boolean,
    parse_date,
    parse_row,
)
from apps.authn.services.import_members.template import generate_template_excel

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - test env should include openpyxl
    load_workbook = None


class GenerateTemplateExcelTests(TestCase):
    def test_template_is_valid_workbook_with_headers_and_instructions(self):
        data = generate_template_excel()
        self.assertIsInstance(data, bytes)
        self.assertGreater(len(data), 0)

        workbook = load_workbook(filename=BytesIO(data))
        self.assertIn("Members", workbook.sheetnames)
        self.assertIn("Instructions", workbook.sheetnames)

        members_sheet = workbook["Members"]
        header_values = [cell.value for cell in members_sheet[1]]
        self.assertIn("Primary Email", header_values)
        self.assertIn("Phone Number", header_values)

        # Example row populated on row 2
        example_first = members_sheet.cell(row=2, column=1).value
        self.assertEqual(example_first, "John")

        instructions = workbook["Instructions"]
        self.assertEqual(instructions.cell(row=1, column=1).value, "Field Descriptions")


class ParsingHelperTests(TestCase):
    def test_generate_random_password_length_and_charset(self):
        pw = generate_random_password(16)
        self.assertEqual(len(pw), 16)

    def test_normalize_header_empty(self):
        self.assertEqual(normalize_header(""), "")

    def test_normalize_header_maps_known_alias(self):
        self.assertEqual(normalize_header("Primary Email"), "primary_email")

    def test_normalize_header_unknown_lowercased(self):
        self.assertEqual(normalize_header("Some Random"), "some random")

    def test_parse_boolean_variants(self):
        self.assertTrue(parse_boolean(True))
        self.assertFalse(parse_boolean(None))
        self.assertTrue(parse_boolean("yes"))
        self.assertFalse(parse_boolean("no"))

    def test_parse_date_none(self):
        self.assertIsNone(parse_date(None))

    def test_parse_date_naive_datetime_made_aware(self):
        result = parse_date(datetime(2024, 1, 15, 10, 30))
        self.assertTrue(timezone.is_aware(result))

    def test_parse_date_aware_datetime_passthrough(self):
        aware = timezone.make_aware(datetime(2024, 1, 15, 10, 30))
        self.assertEqual(parse_date(aware), aware)

    def test_parse_date_empty_string(self):
        self.assertIsNone(parse_date("   "))

    def test_parse_date_recognizes_format(self):
        result = parse_date("2024-01-15")
        self.assertEqual(result.year, 2024)
        self.assertEqual(result.month, 1)
        self.assertEqual(result.day, 15)

    def test_parse_date_unrecognized_returns_none(self):
        self.assertIsNone(parse_date("not-a-date"))

    def test_clean_phone_none(self):
        self.assertIsNone(clean_phone(None))

    def test_clean_phone_empty(self):
        self.assertIsNone(clean_phone("   "))

    def test_clean_phone_strips_trailing_dot_zero(self):
        self.assertEqual(clean_phone("12095551234.0"), "+12095551234")

    def test_clean_phone_no_digits_returns_none(self):
        self.assertIsNone(clean_phone("abc"))

    def test_clean_phone_with_plus(self):
        self.assertEqual(clean_phone("+1 (209) 555-1234"), "+12095551234")

    def test_clean_phone_without_plus(self):
        self.assertEqual(clean_phone("2095551234"), "+2095551234")

    def test_parse_row_full(self):
        row = parse_row(
            5,
            {
                "primary_email": " Main@Example.com ",
                "first_name": " Ada ",
                "last_name": " Lovelace ",
                "secondary_email": "alt@example.com",
                "is_active": "yes",
                "is_staff": "no",
                "when_started": "2024-01-15",
                "phone_number": "+12095551234",
            },
        )
        self.assertEqual(row["row"], 5)
        self.assertEqual(row["primary_email"], "Main@Example.com")
        self.assertEqual(row["first_name"], "Ada")
        self.assertTrue(row["is_active"])
        self.assertFalse(row["is_staff"])
        self.assertEqual(row["phone_number"], "+12095551234")

    def test_parse_row_minimal_defaults(self):
        row = parse_row(2, {})
        self.assertIsNone(row["primary_email"])
        self.assertEqual(row["first_name"], "")
        self.assertIsNone(row["is_active"])
        self.assertIsNone(row["is_staff"])
        self.assertIsNone(row["phone_number"])
