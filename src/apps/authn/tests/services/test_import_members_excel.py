from io import BytesIO
from unittest.mock import patch

from django.test import TestCase

from apps.authn.models import ContactEmail, ContactPhone, Member
from apps.authn.services.import_members import import_members_from_excel

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - test environment should include openpyxl
    Workbook = None


class ImportMembersExcelTests(TestCase):
    HEADERS = ["Primary Email", "First Name", "Last Name", "Organization"]

    def _build_workbook(self, rows: list[list[str]], headers: list[str] | None = None):
        if Workbook is None:
            self.fail("openpyxl is required for import tests")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(headers if headers is not None else self.HEADERS)
        for row in rows:
            worksheet.append(row)

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return payload

    def _build_full_workbook(self, rows: list[list]):
        headers = [
            "Primary Email",
            "First Name",
            "Last Name",
            "Secondary Email",
            "Phone Number",
            "Organization",
        ]
        return self._build_workbook(rows, headers=headers)

    def test_missing_last_name_skips_new_member(self):
        workbook = self._build_workbook(
            [
                ["new-member@example.com", "Ada", "", "Acme"],
            ]
        )

        result = import_members_from_excel(workbook)

        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.skipped_count, 1)
        self.assertIn("Row 2: Missing last name", result.errors)
        self.assertFalse(ContactEmail.objects.filter(email_address="new-member@example.com").exists())

    def test_update_existing_does_not_clear_names_when_import_row_leaves_them_blank(self):
        member = Member.objects.create_user(
            password="StrongPass123!",
            first_name="Existing",
            last_name="Member",
            organization="Original Org",
            is_active=True,
        )
        ContactEmail.objects.create(
            member=member,
            email_address="existing@example.com",
            email_type="primary",
            verified=True,
        )

        workbook = self._build_workbook(
            [
                ["existing@example.com", "", "", "Updated Org"],
            ]
        )

        result = import_members_from_excel(workbook, update_existing=True)

        member.refresh_from_db()
        self.assertTrue(result.success)
        self.assertEqual(result.updated_count, 1)
        self.assertEqual(member.first_name, "Existing")
        self.assertEqual(member.last_name, "Member")
        self.assertEqual(member.organization, "Updated Org")

    # ── File-level guards ────────────────────────────────

    def test_openpyxl_missing_returns_error(self):
        with patch("apps.authn.services.import_members.excel.load_workbook", None):
            result = import_members_from_excel(BytesIO(b""))
        self.assertFalse(result.success)
        self.assertIn("openpyxl library not installed. Please run: pip install openpyxl", result.errors)

    def test_empty_file_returns_error(self):
        if Workbook is None:
            self.fail("openpyxl required")
        workbook = Workbook()
        # delete default rows by leaving sheet empty: iter_rows yields nothing
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        result = import_members_from_excel(payload)
        self.assertFalse(result.success)
        self.assertIn("Excel file is empty", result.errors)

    def test_missing_primary_email_column_returns_error(self):
        workbook = self._build_workbook([["Ada", "Lovelace"]], headers=["First Name", "Last Name"])
        result = import_members_from_excel(workbook)
        self.assertFalse(result.success)
        self.assertIn("Excel file must contain a 'Primary Email' column", result.errors)

    def test_blank_rows_are_skipped(self):
        workbook = self._build_workbook(
            [
                [None, None, None, None],
                ["solo@example.com", "Ada", "Lovelace", "Acme"],
            ]
        )
        result = import_members_from_excel(workbook)
        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 1)

    def test_missing_primary_email_in_row_skips(self):
        workbook = self._build_workbook([["", "Ada", "Lovelace", "Acme"]])
        result = import_members_from_excel(workbook)
        self.assertTrue(result.success)
        self.assertEqual(result.skipped_count, 1)
        self.assertIn("Row 2: Missing primary email", result.errors)

    def test_duplicate_email_within_file_skips_second(self):
        workbook = self._build_workbook(
            [
                ["dup@example.com", "Ada", "Lovelace", "Acme"],
                ["DUP@example.com", "Grace", "Hopper", "Navy"],
            ]
        )
        result = import_members_from_excel(workbook)
        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.skipped_count, 1)
        self.assertTrue(any("Duplicate email in file" in e for e in result.errors))

    def test_no_valid_parsed_rows_returns_early(self):
        workbook = self._build_workbook([["", "Ada", "Lovelace", "Acme"]])
        result = import_members_from_excel(workbook)
        # parsed_rows empty after skip -> returns result with skip recorded
        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 0)

    def test_existing_member_skipped_when_update_disabled(self):
        member = Member.objects.create_user(password="StrongPass123!", first_name="Ex", last_name="Member")
        ContactEmail.objects.create(
            member=member, email_address="exists@example.com", email_type="primary", verified=True
        )
        workbook = self._build_workbook([["exists@example.com", "New", "Name", "NewOrg"]])
        result = import_members_from_excel(workbook, update_existing=False)
        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.skipped_count, 1)
        self.assertTrue(any("already exists" in e for e in result.errors))

    def test_missing_first_name_skips_new_member(self):
        workbook = self._build_workbook([["nofn@example.com", "", "Lovelace", "Acme"]])
        result = import_members_from_excel(workbook)
        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.skipped_count, 1)
        self.assertIn("Row 2: Missing first name", result.errors)

    def test_creates_member_with_emails_and_phone(self):
        workbook = self._build_full_workbook(
            [
                ["main@example.com", "Ada", "Lovelace", "alt@example.com", "+12095551234", "Acme"],
            ]
        )
        result = import_members_from_excel(workbook)
        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 1)
        self.assertTrue(ContactEmail.objects.filter(email_address="main@example.com").exists())
        self.assertTrue(ContactEmail.objects.filter(email_address="alt@example.com").exists())
        self.assertTrue(ContactPhone.objects.filter(phone_number="2095551234").exists())

    def test_exception_during_create_sets_failure(self):
        workbook = self._build_workbook([["boom@example.com", "Ada", "Lovelace", "Acme"]])
        with patch(
            "apps.authn.services.import_members.excel.Member.objects.bulk_create",
            side_effect=RuntimeError("db blew up"),
        ):
            result = import_members_from_excel(workbook)
        self.assertFalse(result.success)
        self.assertEqual(result.created_count, 0)
        self.assertTrue(any("Error creating members" in e for e in result.errors))

    # ── Update path (operations.py) ──────────────────────

    def test_update_existing_updates_contacts_and_phone(self):
        member = Member.objects.create_user(password="StrongPass123!", first_name="Old", last_name="Name")
        ContactEmail.objects.create(
            member=member, email_address="upd@example.com", email_type="primary", verified=False, subscribe=False
        )
        workbook = self._build_full_workbook(
            [
                ["upd@example.com", "New", "Name", "second@example.com", "+12095559999", "NewOrg"],
            ]
        )
        result = import_members_from_excel(workbook, update_existing=True)
        member.refresh_from_db()
        self.assertTrue(result.success)
        self.assertEqual(result.updated_count, 1)
        self.assertEqual(member.first_name, "New")
        self.assertEqual(member.organization, "NewOrg")
        self.assertTrue(ContactEmail.objects.filter(member=member, email_type="secondary").exists())
        self.assertTrue(ContactPhone.objects.filter(member=member, phone_number="2095559999").exists())

    def test_update_existing_replaces_phone_and_clears_secondary(self):
        member = Member.objects.create_user(password="StrongPass123!", first_name="Old", last_name="Name")
        ContactEmail.objects.create(
            member=member, email_address="repl@example.com", email_type="primary", verified=True
        )
        ContactEmail.objects.create(
            member=member, email_address="oldsec@example.com", email_type="secondary", verified=True
        )
        ContactPhone.objects.create(member=member, phone_number="2095550000", region="1-US")
        # Import with no secondary email and a new phone -> secondary deleted, phone replaced
        workbook = self._build_full_workbook(
            [
                ["repl@example.com", "New", "Name", "", "+12095557777", "Org"],
            ]
        )
        result = import_members_from_excel(workbook, update_existing=True)
        member.refresh_from_db()
        self.assertTrue(result.success)
        self.assertEqual(result.updated_count, 1)
        self.assertFalse(ContactEmail.objects.filter(member=member, email_type="secondary").exists())
        self.assertTrue(ContactPhone.objects.filter(member=member, phone_number="2095557777").exists())
        self.assertFalse(ContactPhone.objects.filter(member=member, phone_number="2095550000").exists())

    def test_update_existing_member_not_resolved_is_skipped(self):
        # The primary ContactEmail exists (so the row is routed to update), but the
        # row's email only matches a *secondary* contact -> member_map lookup misses
        # -> bulk_update_members skips it (operations.py lines 30-32).
        member = Member.objects.create_user(password="StrongPass123!", first_name="Old", last_name="Name")
        # Primary contact under a DIFFERENT casing-insensitive address so the
        # existing_emails set sees it (routes to update), but it is stored as secondary.
        ContactEmail.objects.create(
            member=member, email_address="ghost@example.com", email_type="secondary", verified=True
        )
        workbook = self._build_workbook([["ghost@example.com", "New", "Name", "Org"]])

        result = import_members_from_excel(workbook, update_existing=True)

        # No primary-type contact matched -> member_map miss -> skipped.
        self.assertEqual(result.updated_count, 0)
        self.assertEqual(result.skipped_count, 1)

    def test_update_single_member_error_is_recorded(self):
        member = Member.objects.create_user(password="StrongPass123!", first_name="Old", last_name="Name")
        ContactEmail.objects.create(
            member=member, email_address="errrow@example.com", email_type="primary", verified=True
        )
        workbook = self._build_workbook([["errrow@example.com", "New", "Name", "Org"]])

        from apps.authn.services.import_members import operations

        with patch.object(operations, "update_single_member", side_effect=RuntimeError("update boom")):
            result = import_members_from_excel(workbook, update_existing=True)

        self.assertEqual(result.skipped_count, 1)
        self.assertTrue(any("update boom" in e for e in result.errors))
