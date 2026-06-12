"""Regression test: member XLSX export neutralizes spreadsheet formula injection.

openpyxl writes a Python string beginning with ``=`` as a live formula, so a
member who sets their name to ``=HYPERLINK(...)`` would otherwise get a formula
that executes when a staffer opens the export. ``safe_sheet_value`` prefixes a
quote so the value stays literal text.
"""

from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from apps.authn.models import Member
from apps.authn.services.export_members import export_members_to_excel


class ExportMembersFormulaSafetyTest(TestCase):
    def test_export_neutralizes_formula_in_member_fields(self):
        Member.objects.create_user(
            password="StrongPass123!",
            first_name='=HYPERLINK("https://attacker.example","click")',
            last_name="@SUM(A1:A9)",
            organization="-2+3",
            is_active=True,
        )

        content = export_members_to_excel(Member.objects.all())
        worksheet = load_workbook(BytesIO(content)).active

        # No string cell is written as / left as a live formula trigger.
        formula_cells = [
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@"))
        ]
        self.assertEqual(formula_cells, [])

        # The injected name is preserved as quoted literal text.
        values = {cell.value for row in worksheet.iter_rows() for cell in row}
        self.assertIn('\'=HYPERLINK("https://attacker.example","click")', values)
