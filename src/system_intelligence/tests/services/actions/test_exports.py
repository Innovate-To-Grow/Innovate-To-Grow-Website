from openpyxl import load_workbook

from system_intelligence.models import SystemIntelligenceExport
from system_intelligence.services import actions
from system_intelligence.services.exports import EXPORT_HARD_LIMIT, ExportError, create_export

from .base import SystemIntelligenceActionBase


class CreateExportTests(SystemIntelligenceActionBase):
    def _make_members(self, count, prefix="User"):
        from authn.models import Member

        for index in range(count):
            Member.objects.create_user(
                email=f"{prefix.lower()}{index}@example.com",
                password="testpass123",
                first_name=prefix,
                last_name=str(index),
            )

    def test_export_writes_workbook_with_safe_columns_and_row_count(self):
        self._make_members(3)
        export = create_export(app_label="authn", model_name="Member")
        self.assertIsInstance(export, SystemIntelligenceExport)
        self.assertEqual(export.row_count, 4)  # 3 + the admin user
        self.assertEqual(export.model_label, "authn.Member")
        self.assertTrue(export.filename.endswith(".xlsx"))
        # All admin users + the 3 we created should be in the workbook.
        workbook = load_workbook(export.file.path, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(len(rows), 5)  # header + 4 data rows
        headers = rows[0]
        self.assertIn("First Name", headers)
        # Sensitive fields must never appear in the export header row.
        self.assertNotIn("password", [h.lower() if isinstance(h, str) else "" for h in headers])

    def test_export_respects_filters_and_limit(self):
        self._make_members(3, prefix="Alpha")
        self._make_members(2, prefix="Beta")
        export = create_export(
            app_label="authn",
            model_name="Member",
            filters={"first_name__iexact": "Alpha"},
            limit=2,
        )
        self.assertEqual(export.row_count, 2)

    def test_export_uses_explicit_field_selection(self):
        self._make_members(1, prefix="Just")
        export = create_export(
            app_label="authn",
            model_name="Member",
            fields=["first_name", "email"],
        )
        self.assertEqual(set(export.field_names), {"first_name", "email"})
        workbook = load_workbook(export.file.path, read_only=True)
        sheet = workbook.active
        headers = next(sheet.iter_rows(values_only=True))
        self.assertEqual(len(headers), 2)

    def test_export_raises_when_no_rows_match(self):
        with self.assertRaises(ExportError) as cm:
            create_export(
                app_label="authn",
                model_name="Member",
                filters={"first_name": "ZZZZ-no-such-name"},
            )
        self.assertIn("No rows match", str(cm.exception))

    def test_export_rejects_denied_model(self):
        with self.assertRaises(ExportError):
            create_export(app_label="auth", model_name="Permission")

    def test_export_rejects_unknown_field(self):
        with self.assertRaises(ExportError):
            create_export(
                app_label="authn",
                model_name="Member",
                fields=["email", "definitely_not_a_field"],
            )

    def test_export_rejects_unknown_filter_lookup(self):
        with self.assertRaises(ExportError):
            create_export(
                app_label="authn",
                model_name="Member",
                filters={"email__regex": ".*"},
            )

    def test_export_caps_limit_at_hard_max(self):
        export = create_export(
            app_label="authn",
            model_name="Member",
            limit=EXPORT_HARD_LIMIT * 10,
        )
        # Only the admin user exists, so row_count is 1; the cap was applied without an error.
        self.assertEqual(export.row_count, 1)

    def test_export_requires_active_conversation(self):
        actions.reset_action_context(self.context_tokens)
        try:
            with self.assertRaises(Exception) as cm:  # ActionRequestError surfaces as raised; we catch broadly.
                create_export(app_label="authn", model_name="Member")
            self.assertIn("conversation", str(cm.exception).lower())
        finally:
            self.context_tokens = actions.set_action_context(str(self.conversation.pk), str(self.admin_user.pk))
