import io

import openpyxl
from django.core.cache import cache
from django.test import TestCase

from projects.models import Project, Semester
from projects.services.import_excel import import_projects_from_excel


def _make_excel(semesters_data, projects_data):
    """Create an in-memory Excel workbook with Semesters and Projects sheets."""
    wb = openpyxl.Workbook()

    ws_sem = wb.active
    ws_sem.title = "Semesters"
    ws_sem.append(["year", "season", "label"])
    for row in semesters_data:
        ws_sem.append(row)

    ws_proj = wb.create_sheet("Projects")
    ws_proj.append(
        [
            "year",
            "season",
            "class_code",
            "team_number",
            "team_name",
            "project_title",
            "organization",
            "industry",
            "abstract",
            "student_names",
            "track",
            "presentation_order",
        ]
    )
    for row in projects_data:
        ws_proj.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ImportExcelServiceTest(TestCase):
    # noinspection PyPep8Naming,PyMethodMayBeStatic
    def setUp(self):
        cache.clear()

    def test_imports_semesters_and_projects(self):
        excel = _make_excel(
            semesters_data=[[2025, 1, "2025-1 Spring"]],
            projects_data=[
                [2025, 1, "CAP", "101", "Team A", "Project Alpha", "Org A", "Tech", "Abstract text", "Alice, Bob", 1, 2]
            ],
        )

        stats = import_projects_from_excel(excel)

        self.assertEqual(stats["semesters_created"], 1)
        self.assertEqual(stats["projects_created"], 1)
        self.assertEqual(Semester.objects.count(), 1)
        self.assertEqual(Project.objects.count(), 1)

        project = Project.objects.first()
        self.assertEqual(project.project_title, "Project Alpha")
        self.assertEqual(project.team_number, "101")
        self.assertEqual(project.organization, "Org A")
        self.assertEqual(project.track, 1)
        self.assertEqual(project.presentation_order, 2)

    def test_updates_existing_projects(self):
        excel1 = _make_excel(
            semesters_data=[[2025, 1, "2025-1 Spring"]],
            projects_data=[
                [2025, 1, "CAP", "101", "Team A", "Project Alpha", "Org A", "Tech", "Old abstract", "Alice", 1, 1]
            ],
        )
        import_projects_from_excel(excel1)

        excel2 = _make_excel(
            semesters_data=[[2025, 1, "2025-1 Spring"]],
            projects_data=[
                [
                    2025,
                    1,
                    "CAP",
                    "101",
                    "Team A Updated",
                    "Project Alpha",
                    "Org B",
                    "Tech",
                    "New abstract",
                    "Alice, Bob",
                    1,
                    1,
                ]
            ],
        )
        stats = import_projects_from_excel(excel2)

        self.assertEqual(stats["projects_updated"], 1)
        self.assertEqual(stats["projects_created"], 0)
        self.assertEqual(Project.objects.count(), 1)

        project = Project.objects.first()
        self.assertEqual(project.team_name, "Team A Updated")
        self.assertEqual(project.organization, "Org B")
        self.assertEqual(project.abstract, "New abstract")

    def test_skips_rows_without_title(self):
        excel = _make_excel(
            semesters_data=[[2025, 1, "2025-1 Spring"]],
            projects_data=[
                [2025, 1, "CAP", "101", "Team A", "Valid Project", "Org", "", "", "", None, None],
                [2025, 1, "CAP", "102", "Team B", "", "Org", "", "", "", None, None],
                [2025, 1, "CAP", "103", "Team C", None, "Org", "", "", "", None, None],
            ],
        )

        stats = import_projects_from_excel(excel)

        self.assertEqual(stats["projects_created"], 1)
        self.assertEqual(stats["rows_skipped"], 2)

    def test_existing_semester_not_duplicated(self):
        Semester.objects.create(year=2025, season=1)

        excel = _make_excel(
            semesters_data=[[2025, 1, "2025-1 Spring"]],
            projects_data=[[2025, 1, "CAP", "101", "Team A", "Project A", "", "", "", "", None, None]],
        )

        stats = import_projects_from_excel(excel)

        self.assertEqual(stats["semesters_existing"], 1)
        self.assertEqual(stats["semesters_created"], 0)
        self.assertEqual(Semester.objects.count(), 1)

    def test_raises_on_missing_semesters_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = "WrongName"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with self.assertRaises(ValueError) as ctx:
            import_projects_from_excel(buf)
        self.assertIn("Semesters", str(ctx.exception))

    def test_raises_on_missing_projects_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Semesters"
        ws = wb.active
        ws.append(["year", "season", "label"])
        ws.append([2025, 1, "2025-1 Spring"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with self.assertRaises(ValueError) as ctx:
            import_projects_from_excel(buf)
        self.assertIn("Projects", str(ctx.exception))

    def test_handles_none_industry_and_track(self):
        excel = _make_excel(
            semesters_data=[[2020, 2, "2020-2 Fall"]],
            projects_data=[
                [2020, 2, "CAP", "CAP01", "Team X", "Some Project", "Org", None, "Abstract", "Students", None, None]
            ],
        )

        stats = import_projects_from_excel(excel)

        self.assertEqual(stats["projects_created"], 1)
        project = Project.objects.first()
        self.assertEqual(project.industry, "")
        self.assertIsNone(project.track)
        self.assertIsNone(project.presentation_order)
