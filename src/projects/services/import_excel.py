import openpyxl
from django.core.cache import cache

from projects.models import Project, Semester

SEMESTER_HEADERS = {"year", "season", "label"}
PROJECT_HEADERS = {
    "year",
    "season",
    "class_code",
    "team_number",
    "team_name",
    "project_title",
    "organization",
    "industry",
    "abstract",
    "student_names",
    "track",
    "presentation_order",
}


def _to_int(val):
    """Convert a value to int, return None if not possible."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _to_str(val):
    """Convert a value to stripped string, return '' if None."""
    if val is None:
        return ""
    return str(val).strip()


def import_projects_from_excel(file):
    """
    Import semesters and projects from an uploaded Excel file.

    Expected format:
      - Sheet "Semesters": columns year, season, label
      - Sheet "Projects": columns year, season, class_code, team_number,
        team_name, project_title, organization, industry, abstract,
        student_names, track, presentation_order

    Returns dict with counts of created/updated records.
    """
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

    stats = {
        "semesters_created": 0,
        "semesters_existing": 0,
        "projects_created": 0,
        "projects_updated": 0,
        "rows_skipped": 0,
    }

    # --- Semesters ---
    if "Semesters" not in wb.sheetnames:
        wb.close()
        raise ValueError("Excel file must contain a 'Semesters' sheet.")

    ws_sem = wb["Semesters"]
    sem_rows = list(ws_sem.iter_rows(values_only=True))

    if not sem_rows:
        wb.close()
        raise ValueError("Semesters sheet is empty.")

    sem_headers = [str(h).strip().lower() if h else "" for h in sem_rows[0]]
    if not SEMESTER_HEADERS.issubset(set(sem_headers)):
        wb.close()
        raise ValueError(f"Semesters sheet must have columns: {', '.join(sorted(SEMESTER_HEADERS))}")

    sem_col = {name: idx for idx, name in enumerate(sem_headers)}
    semester_cache = {}

    for row in sem_rows[1:]:
        year = _to_int(row[sem_col["year"]])
        season = _to_int(row[sem_col["season"]])
        if not year or not season:
            continue

        semester, created = Semester.objects.get_or_create(year=year, season=season)
        if not semester.is_published:
            semester.is_published = True
            semester.save(update_fields=["is_published"])
        semester_cache[(year, season)] = semester
        if created:
            stats["semesters_created"] += 1
        else:
            stats["semesters_existing"] += 1

    # --- Projects ---
    if "Projects" not in wb.sheetnames:
        wb.close()
        raise ValueError("Excel file must contain a 'Projects' sheet.")

    ws_proj = wb["Projects"]
    proj_rows = list(ws_proj.iter_rows(values_only=True))

    if not proj_rows:
        wb.close()
        raise ValueError("Projects sheet is empty.")

    proj_headers = [str(h).strip().lower() if h else "" for h in proj_rows[0]]
    required = {"year", "season", "project_title"}
    if not required.issubset(set(proj_headers)):
        wb.close()
        raise ValueError(f"Projects sheet must have columns: {', '.join(sorted(required))}")

    proj_col = {name: idx for idx, name in enumerate(proj_headers) if name}

    def get(row, col_name, default=""):
        idx = proj_col.get(col_name)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    for row in proj_rows[1:]:
        year = _to_int(get(row, "year"))
        season = _to_int(get(row, "season"))
        title = _to_str(get(row, "project_title"))

        if not year or not season or not title:
            stats["rows_skipped"] += 1
            continue

        semester = semester_cache.get((year, season))
        if not semester:
            semester, created = Semester.objects.get_or_create(year=year, season=season)
            if not semester.is_published:
                semester.is_published = True
                semester.save(update_fields=["is_published"])
            semester_cache[(year, season)] = semester
            if created:
                stats["semesters_created"] += 1

        team_number = _to_str(get(row, "team_number"))
        defaults = {
            "team_name": _to_str(get(row, "team_name")),
            "project_title": title,
            "organization": _to_str(get(row, "organization")),
            "industry": _to_str(get(row, "industry")),
            "abstract": _to_str(get(row, "abstract")),
            "student_names": _to_str(get(row, "student_names")),
            "track": _to_int(get(row, "track")),
            "presentation_order": _to_int(get(row, "presentation_order")),
            "class_code": _to_str(get(row, "class_code")),
        }

        _, created = Project.objects.update_or_create(
            semester=semester,
            team_number=team_number,
            project_title=title,
            defaults=defaults,
        )
        if created:
            stats["projects_created"] += 1
        else:
            stats["projects_updated"] += 1

    wb.close()

    # Clear cached project data
    cache.delete("projects:current")

    return stats
